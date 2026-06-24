from __future__ import annotations

import shutil
import tempfile
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

_PREFERRED_SHEETS = ["Model vUpd", "Model", "Model v1", "Model vOld"]


def _load_ws(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    for name in _PREFERRED_SHEETS:
        if name in wb.sheetnames:
            return wb, wb[name]
    # fallback: first sheet containing 'Net IRR' in first 50 rows
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(max_row=50):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().upper() == "NET IRR":
                    return wb, ws
    return wb, wb[wb.sheetnames[0]]


def read_model_outputs(path: str) -> dict:
    """
    Extract key deal model outputs from an Excel model file.

    Returns a dict with these keys (all float, None if not found):
      gross_bid   — Gross Bid Price as decimal (e.g. 0.83 for 83c)
      eff_bid     — Effective Bid Price as decimal
      base_irr    — Base Case Secondary Net IRR as decimal
      base_moic   — Base Case Secondary Net MOIC
      bear_irr    — Bear Case Secondary Net IRR as decimal (Downside)
      mgr_irr     — Manager Case Secondary Net IRR as decimal (Upside)
    """
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to read Excel models.")

    tmp = Path(tempfile.mktemp(suffix=".xlsx"))
    shutil.copy2(path, str(tmp))
    try:
        wb, ws = _load_ws(str(tmp))
        cells = {}
        try:
            for row in ws.iter_rows():
                for cell in row:
                    if hasattr(cell, "row") and hasattr(cell, "column"):
                        cells[(cell.row, cell.column)] = cell.value
        finally:
            wb.close()
    finally:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass

    result: dict = {}

    # --- Bid prices ---
    for label, key in [("Gross Bid Price", "gross_bid"), ("Effective Bid Price", "eff_bid")]:
        for (r, c), v in cells.items():
            if isinstance(v, str) and v.strip() == label:
                adj = cells.get((r, c + 1))
                if isinstance(adj, (int, float)):
                    result[key] = float(adj)
                break

    # --- Returns table ---
    # Find the row with the most 'Sec' sub-headers (marks base/bear/mgr secondary columns)
    sec_by_row: dict = {}
    for (r, c), v in cells.items():
        if isinstance(v, str) and v.strip() == "Sec":
            sec_by_row.setdefault(r, []).append(c)

    sec_row: Optional[int] = None
    if sec_by_row:
        sec_row = max(sec_by_row, key=lambda r: len(sec_by_row[r]))

    if sec_row is not None:
        case_row = sec_row - 1

        def case_for_col(sc: int) -> Optional[str]:
            # Look leftward on case_row for the nearest BASE/BEAR/MANAGER header
            for c in range(sc, max(1, sc - 10) - 1, -1):
                v = cells.get((case_row, c))
                if isinstance(v, str):
                    u = v.strip().upper()
                    if "BASE" in u:
                        return "base"
                    if "BEAR" in u:
                        return "bear"
                    if "MANAGER" in u or "MGR" in u or "UPSIDE" in u:
                        return "mgr"
            return None

        case_sec_col: dict = {}
        for sc in sorted(sec_by_row[sec_row]):
            case = case_for_col(sc)
            if case and case not in case_sec_col:
                case_sec_col[case] = sc

        irr_row = moic_row = None
        for r in range(sec_row + 1, sec_row + 15):
            for c in range(1, 6):
                v = cells.get((r, c))
                if isinstance(v, str):
                    u = v.strip().upper()
                    if u == "NET IRR" and irr_row is None:
                        irr_row = r
                    elif u == "NET MOIC" and moic_row is None:
                        moic_row = r

        if irr_row:
            for case, col in case_sec_col.items():
                v = cells.get((irr_row, col))
                if isinstance(v, (int, float)):
                    result[f"{case}_irr"] = float(v)
        if moic_row:
            for case, col in case_sec_col.items():
                v = cells.get((moic_row, col))
                if isinstance(v, (int, float)):
                    result[f"{case}_moic"] = float(v)

    else:
        # Fallback: no Sec sub-headers — read directly from case header columns
        # Find rows with case headers and Net IRR/MOIC rows
        case_header_rows: dict = {}
        for (r, c), v in cells.items():
            if isinstance(v, str):
                u = v.strip().upper()
                if "BASE CASE" in u or u == "BASE":
                    case_header_rows.setdefault(r, {})["base"] = c
                elif "BEAR CASE" in u or u == "BEAR":
                    case_header_rows.setdefault(r, {})["bear"] = c
                elif "MANAGER CASE" in u or "MGR CASE" in u:
                    case_header_rows.setdefault(r, {})["mgr"] = c

        if case_header_rows:
            hdr_row = max(case_header_rows, key=lambda r: len(case_header_rows[r]))
            case_cols = case_header_rows[hdr_row]

            irr_row = moic_row = None
            for r in range(hdr_row + 1, hdr_row + 20):
                for c in range(1, 6):
                    v = cells.get((r, c))
                    if isinstance(v, str):
                        u = v.strip().upper()
                        if u == "NET IRR" and irr_row is None:
                            irr_row = r
                        elif u == "NET MOIC" and moic_row is None:
                            moic_row = r

            if irr_row:
                for case, col in case_cols.items():
                    v = cells.get((irr_row, col + 1))
                    if isinstance(v, (int, float)):
                        result[f"{case}_irr"] = float(v)
            if moic_row:
                for case, col in case_cols.items():
                    v = cells.get((moic_row, col + 1))
                    if isinstance(v, (int, float)):
                        result[f"{case}_moic"] = float(v)

    return result


def read_cashflows(path: str) -> dict:
    """
    Read the Combined Cashflows table from the first sheet.

    Returns:
      {
        'header': {
          'base_irr_sec', 'base_irr_fund', 'bear_irr_sec', 'bear_irr_fund',
          'mgr_irr_sec',  'mgr_irr_fund',
          'base_moic_sec', ... (same pattern),
          'base_dur_sec',  ... (same pattern),
        },
        'rows': [
          {'date': 'Sep-25', 'base_s': float|None, 'base_f': float|None,
           'bear_s': ..., 'bear_f': ..., 'mgr_s': ..., 'mgr_f': ...},
          ...  (only rows where at least one column has abs >= 0.01)
        ]
      }
    """
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to read Excel models.")

    import re as _re
    from datetime import datetime as _dt

    tmp = Path(tempfile.mktemp(suffix=".xlsx"))
    shutil.copy2(path, str(tmp))
    try:
        wb = openpyxl.load_workbook(str(tmp), data_only=True)
        ws = wb[wb.sheetnames[0]]  # first sheet always has bid info + combined cashflows
        cells: dict = {}
        try:
            for row in ws.iter_rows():
                for cell in row:
                    if hasattr(cell, "row") and hasattr(cell, "column"):
                        cells[(cell.row, cell.column)] = cell.value
        finally:
            wb.close()
    finally:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass

    def _fv(v):
        return float(v) if isinstance(v, (int, float)) else None

    # Locate quarterly date rows in column 2 (stored as datetime or "Sep-25" strings)
    _date_pat = _re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2}$")
    date_rows: list = []
    for (r, c), v in sorted(cells.items()):
        if c != 2:
            continue
        if isinstance(v, _dt):
            date_rows.append((r, v.strftime("%b-%y")))
        elif isinstance(v, str) and _date_pat.match(v.strip()):
            date_rows.append((r, v.strip()))
    if not date_rows:
        return {}
    date_rows.sort()
    first_date_row = date_rows[0][0]

    # The S CFs / F CFs sub-header row sits directly above the first date row
    scf_row = first_date_row - 1

    # Find the case header row by scanning upward
    case_cols: dict = {}
    case_row: Optional[int] = None
    for r in range(scf_row - 1, max(1, scf_row - 15) - 1, -1):
        for c in range(1, 20):
            v = cells.get((r, c))
            if isinstance(v, str):
                u = v.strip().upper()
                if "BASE" in u and "CASE" in u and "base" not in case_cols:
                    case_cols["base"] = c
                    case_row = r
                elif "BEAR" in u and "CASE" in u and "bear" not in case_cols:
                    case_cols["bear"] = c
                elif ("MANAGER" in u or "MGR" in u) and "mgr" not in case_cols:
                    case_cols["mgr"] = c
        if case_cols:
            break

    if not case_cols or case_row is None:
        return {}

    # For each case column, find the F CFs column (first col to the right whose label starts with "F")
    # Note: "F CFs".upper() = "F CFS" which contains "S", so we cannot use "S" not in check —
    # use startswith("F") to distinguish "F CFs" from "S CFs" reliably.
    def get_f_col(start_c: int) -> int:
        for offset in range(1, 6):
            v = cells.get((scf_row, start_c + offset))
            if isinstance(v, str) and v.strip().upper().startswith("F"):
                return start_c + offset
        return start_c + 2  # fallback

    base_s = case_cols.get("base", 3)
    bear_s = case_cols.get("bear", 6)
    mgr_s  = case_cols.get("mgr",  9)
    base_f = get_f_col(base_s)
    bear_f = get_f_col(bear_s)
    mgr_f  = get_f_col(mgr_s)

    # Read Net IRR, Net MOIC, Duration header rows between case_row and scf_row
    h: dict = {}
    for r in range(case_row + 1, scf_row):
        label = cells.get((r, 2))
        if not isinstance(label, str):
            continue
        lu = label.strip().upper()
        if "NET IRR" in lu:
            h.update({
                "base_irr_sec":  _fv(cells.get((r, base_s))),
                "base_irr_fund": _fv(cells.get((r, base_f))),
                "bear_irr_sec":  _fv(cells.get((r, bear_s))),
                "bear_irr_fund": _fv(cells.get((r, bear_f))),
                "mgr_irr_sec":   _fv(cells.get((r, mgr_s))),
                "mgr_irr_fund":  _fv(cells.get((r, mgr_f))),
            })
        elif "NET MOIC" in lu:
            h.update({
                "base_moic_sec":  _fv(cells.get((r, base_s))),
                "base_moic_fund": _fv(cells.get((r, base_f))),
                "bear_moic_sec":  _fv(cells.get((r, bear_s))),
                "bear_moic_fund": _fv(cells.get((r, bear_f))),
                "mgr_moic_sec":   _fv(cells.get((r, mgr_s))),
                "mgr_moic_fund":  _fv(cells.get((r, mgr_f))),
            })
        elif "DURATION" in lu:
            h.update({
                "base_dur_sec":  _fv(cells.get((r, base_s))),
                "base_dur_fund": _fv(cells.get((r, base_f))),
                "bear_dur_sec":  _fv(cells.get((r, bear_s))),
                "bear_dur_fund": _fv(cells.get((r, bear_f))),
                "mgr_dur_sec":   _fv(cells.get((r, mgr_s))),
                "mgr_dur_fund":  _fv(cells.get((r, mgr_f))),
            })

    # Read cashflow data rows; omit trailing rows where all six values are near-zero
    THRESHOLD = 0.01
    value_cols = [base_s, base_f, bear_s, bear_f, mgr_s, mgr_f]
    keys = ["base_s", "base_f", "bear_s", "bear_f", "mgr_s", "mgr_f"]

    cf_rows = []
    for r, date_str in date_rows:
        vals = [_fv(cells.get((r, c))) for c in value_cols]
        if any(v is not None and abs(v) >= THRESHOLD for v in vals):
            entry = {"date": date_str}
            for k, v in zip(keys, vals):
                entry[k] = v if (v is not None and abs(v) >= THRESHOLD) else None
            cf_rows.append(entry)

    return {"header": h, "rows": cf_rows}
