from __future__ import annotations

import io
import json
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


DEFAULT_FIELD_SYNONYMS: Dict[str, List[str]] = {
    "investment_name": [
        "investment name",
        "investment",
        "company",
        "issuer",
        "borrower",
        "asset",
        "position",
        "name",
    ],
    "record_date_nav": [
        "record date nav",
        "rd nav",
        "nav",
        "net asset value",
        "nav amount",
        "fund nav",
        "record date nav ($m)",
        "record date nav ($)",
    ],
    "asset_class": [
        "asset class",
        "asset type",
        "primary asset class",
        "bucket",
    ],
    "security_type": [
        "security type",
        "security",
        "security class",
        "instrument type",
    ],
    "geography": [
        "geography",
        "region",
        "country",
        "location",
        "geo",
    ],
    "sub_asset_class": [
        "sub asset class",
        "sub-asset class",
        "subasset class",
        "strategy",
        "strategy bucket",
    ],
}


def canonicalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


# Common legal-entity suffix tokens dropped when matching the same issuer across funds.
_LEGAL_SUFFIX_TOKENS = {
    "inc", "incorporated", "llc", "ltd", "limited", "lp", "llp", "plc", "corp",
    "corporation", "co", "company", "holdings", "holding", "group", "partners",
    "sa", "ag", "gmbh", "nv", "bv", "sarl", "spa", "pte", "ab", "as", "oy",
}


def canonical_issuer(name: Any) -> str:
    """Canonical key for matching the same issuer across funds despite slight naming
    differences: drop case, punctuation, instrument suffixes (e.g. '– Sr. Sub Debt'),
    trailing parentheticals (e.g. '(fka ...)', '(ii)') and legal suffixes.

    Order: strip instrument first, THEN peel parens — this correctly handles names
    like 'CREO Group (fka Nursery Supplies) – Sr. Sub Debt' where the paren is in
    the middle until after the instrument is removed."""
    text = str(name or "").strip().lower()
    # 1. Strip instrument descriptor after spaced dash first
    text = re.split(r"\s[–—-]\s", text, maxsplit=1)[0].strip()
    # 2. Peel trailing parentheticals (loop handles multiple, e.g. "(fka X) (ii)")
    prev = None
    while prev != text:
        prev = text
        text = re.sub(r"\s*\([^()]*\)\s*$", "", text).strip()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    tokens = [t for t in text.split() if t]
    while tokens and tokens[-1] in _LEGAL_SUFFIX_TOKENS:
        tokens.pop()
    return " ".join(tokens)


def parse_force_splits(text: Any) -> set:
    """Parse a textarea of raw names (one per line) that should NOT be merged with
    other issuers even if they share the same canonical key."""
    out = set()
    for line in str(text or "").splitlines():
        name = line.strip().lower()
        if name:
            out.add(name)
    return out


_PAREN_TAIL = re.compile(r"\s*\([^()]*\)\s*$")


def display_issuer(name: Any) -> str:
    """Human-readable issuer name: drop trailing parentheticals and the instrument
    descriptor after a spaced dash, but keep case/legal form ("GoHealth Inc.")."""
    s = str(name or "").strip()
    prev = None
    while prev != s:
        prev = s
        s = _PAREN_TAIL.sub("", s).strip()
    return re.split(r"\s[–—-]\s", s, maxsplit=1)[0].strip()


def _shortest_display(variants) -> str:
    cands = [c for c in (display_issuer(v) for v in variants) if c]
    return min(cands, key=lambda n: (len(n), n.lower())) if cands else ""


def parse_issuer_aliases(text: Any) -> Dict[str, str]:
    """Parse alias lines 'variant(s) => issuer' into {canonical(variant): canonical(issuer)}.
    The left side may list several variants separated by ; or |. Lets the user force
    a merge the auto-normalizer can't safely make (e.g. 'Noble => Noble Supply and Logistics')."""
    out: Dict[str, str] = {}
    if not text:
        return out
    for line in str(text).splitlines():
        raw = line.strip()
        if not raw or raw.startswith("#"):
            continue
        parts = re.split(r"\s*(?:=>|->|=)\s*", raw, maxsplit=1)
        if len(parts) != 2:
            continue
        target = canonical_issuer(parts[1])
        if not target:
            continue
        for variant in re.split(r"[;|]", parts[0]):
            key = canonical_issuer(variant)
            if key:
                out[key] = target
    return out


def _effective_key(name: Any, alias_map: Dict[str, str], split_set: set = None) -> str:
    raw_lower = str(name or "").strip().lower()
    # Force-split: bypass canonical matching, give this name its own unique key
    if split_set and raw_lower in split_set:
        return f"__split__{raw_lower}"
    key = canonical_issuer(name) or raw_lower
    return alias_map.get(key, key) if alias_map else key


def parse_mapping_rules(text: str) -> Dict[str, str]:
    rules: Dict[str, str] = {}
    for line in (text or "").splitlines():
        raw = line.strip()
        if not raw or raw.startswith("#"):
            continue
        match = re.split(r"\s*(?:=>|->|=)\s*", raw, maxsplit=1)
        if len(match) != 2:
            continue
        left, right = match[0].strip(), match[1].strip()
        if left and right:
            rules[canonicalize_text(left)] = right
    return rules


def parse_manual_allocation(text: str) -> Dict[str, float]:
    allocations: Dict[str, float] = {}
    for raw_line in (text or "").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        parts = re.split(r"[;,]", line)
        if len(parts) == 1:
            parts = [line]
        for part in parts:
            chunk = part.strip()
            if not chunk:
                continue
            match = re.split(r"\s*(?:=>|->|=)\s*", chunk, maxsplit=1)
            if len(match) == 2:
                label, value = match[0].strip(), match[1].strip()
            else:
                pieces = chunk.rsplit(" ", 1)
                if len(pieces) != 2:
                    continue
                label, value = pieces[0].strip(), pieces[1].strip()
            if not label or not value:
                continue
            num = value.replace("%", "").strip()
            try:
                allocations[label] = float(num) / 100.0 if "%" in value or float(num) > 1 else float(num)
            except ValueError:
                continue
    total = sum(max(v, 0.0) for v in allocations.values())
    if total > 0 and abs(total - 1.0) > 1e-6:
        allocations = {label: max(v, 0.0) / total for label, v in allocations.items()}
    return allocations


def parse_manual_family_text(text: str) -> Dict[str, List[Tuple[str, float]]]:
    families: Dict[str, List[Tuple[str, float]]] = {}
    if not text:
        return families
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        match = re.match(
            r"^(geography|asset\s*class|asset_class|security\s*type|security_type|sub\s*asset\s*class|sub_asset_class)\s*[:=]\s*(.+)$",
            line,
            flags=re.IGNORECASE,
        )
        if match:
            family_name = match.group(1)
            payload = match.group(2)
        else:
            family_name = "geography"
            payload = line
        family = normalize_family_name(family_name)
        items: List[Tuple[str, float]] = []
        for label, value in parse_manual_allocation(payload).items():
            items.append((label, value))
        if items:
            total = sum(v for _, v in items)
            if total > 0 and abs(total - 1.0) > 1e-6:
                items = [(label, value / total) for label, value in items]
            families[family] = items
    return families


def normalize_family_name(name: Any) -> str:
    family = canonicalize_text(name).replace(" ", "_")
    aliases = {
        "geographies": "geography",
        "geo": "geography",
        "asset": "asset_class",
        "assetclass": "asset_class",
        "asset_classes": "asset_class",
        "security": "security_type",
        "security_types": "security_type",
        "subassetclass": "sub_asset_class",
    }
    return aliases.get(family, family)


def parse_manual_overrides(value: Any) -> Dict[str, List[Tuple[str, float]]]:
    if not value:
        return {}
    if isinstance(value, dict):
        families: Dict[str, List[Tuple[str, float]]] = {}
        for family_name, items in value.items():
            family = normalize_family_name(family_name)
            parsed_items: List[Tuple[str, float]] = []
            if isinstance(items, dict):
                items = [
                    {"label": label, "pct": pct}
                    for label, pct in items.items()
                ]
            if isinstance(items, list):
                for item in items:
                    if isinstance(item, dict):
                        label = item.get("label") or item.get("name") or item.get("value")
                        pct = item.get("pct", item.get("percentage", item.get("value")))
                    elif isinstance(item, (list, tuple)) and len(item) >= 2:
                        label, pct = item[0], item[1]
                    else:
                        continue
                    if label is None or pct is None:
                        continue
                    try:
                        parsed_items.append((str(label), float(pct)))
                    except (TypeError, ValueError):
                        continue
            if parsed_items:
                total = sum(v for _, v in parsed_items)
                if total > 0 and abs(total - 1.0) > 1e-6:
                    parsed_items = [(label, value / total) for label, value in parsed_items]
                families[family] = parsed_items
        return families
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return {}
        if stripped.startswith("{"):
            try:
                decoded = json.loads(stripped)
            except json.JSONDecodeError:
                decoded = None
            if isinstance(decoded, dict):
                return parse_manual_overrides(decoded)
        return parse_manual_family_text(stripped)
    return {}


def normalize_category(value: Any, rules: Dict[str, str]) -> str:
    if value is None:
        return "Unclassified"
    if pd.isna(value):
        return "Unclassified"
    text = str(value).strip()
    if not text:
        return "Unclassified"
    key = canonicalize_text(text)
    if key in rules:
        return rules[key]
    return text


def to_numeric_series(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(r"[\$,]", "", regex=True)
        .str.replace(r"\(([^)]+)\)", r"-\1", regex=True)
        .str.replace(r"[^0-9.\-]", "", regex=True)
        .replace({"": None, "nan": None, "None": None})
    )
    return pd.to_numeric(cleaned, errors="coerce")


def detect_field_columns(columns: Iterable[Any]) -> Dict[str, List[str]]:
    cols = [str(c) for c in columns]
    ranked: Dict[str, List[str]] = {}
    for field, synonyms in DEFAULT_FIELD_SYNONYMS.items():
        scored: List[Tuple[int, str]] = []
        for col in cols:
            score = _score_column(col, synonyms)
            if score > 0:
                scored.append((score, col))
        scored.sort(key=lambda item: (-item[0], item[1].lower()))
        ranked[field] = [col for _, col in scored[:5]]
    return ranked


def _score_column(column: str, synonyms: List[str]) -> int:
    canon_col = canonicalize_text(column)
    if not canon_col:
        return 0
    col_tokens = set(canon_col.split())
    best = 0
    for syn in synonyms:
        canon_syn = canonicalize_text(syn)
        syn_tokens = set(canon_syn.split())
        if canon_col == canon_syn:
            best = max(best, 100)
        elif canon_syn and (canon_syn in canon_col or canon_col in canon_syn):
            best = max(best, 90)
        overlap = len(col_tokens & syn_tokens)
        if overlap:
            best = max(best, 60 + overlap * 8)
    return best


def _is_generic_column_name(value: Any) -> bool:
    text = canonicalize_text(value)
    if not text:
        return True
    if text.startswith("unnamed"):
        return True
    if text in {"x", "y", "z"}:
        return True
    if text.isdigit():
        return True
    if re.fullmatch(r"column \d+", text):
        return True
    return False


def _looks_numeric_label(value: Any) -> bool:
    text = canonicalize_text(value)
    if not text:
        return False
    if text.startswith("column "):
        return False
    return bool(re.fullmatch(r"[-+]?\d+(\.\d+)?([eE][-+]?\d+)?", text))


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    cleaned = df.copy()
    cleaned = cleaned.dropna(axis=0, how="all")
    cleaned = cleaned.dropna(axis=1, how="all")
    cleaned = cleaned.reset_index(drop=True)
    return cleaned


def _prepare_headerless_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = _clean_dataframe(df)
    cleaned.columns = [f"Column {idx + 1}" for idx in range(len(cleaned.columns))]
    return cleaned


def _detect_sheet_mode(header_df: pd.DataFrame, data_df: pd.DataFrame) -> str:
    header_df = _clean_dataframe(header_df)
    data_df = _prepare_headerless_dataframe(data_df)
    if header_df.empty and not data_df.empty:
        return "data"
    if data_df.empty and not header_df.empty:
        return "header"
    if header_df.empty and data_df.empty:
        return "header"

    header_columns = [str(c) for c in header_df.columns]
    generic_count = sum(_is_generic_column_name(c) for c in header_columns)
    numeric_like_count = sum(_looks_numeric_label(c) for c in header_columns)
    synonym_matches = detect_field_columns(header_df.columns)
    match_count = sum(1 for values in synonym_matches.values() if values)
    textish_columns = sum(
        1
        for col in header_df.columns
        if _score_column_from_data(pd.Series(header_df[col])).get("text_ratio", 0.0) > 0.7
    )
    generic_ratio = generic_count / max(len(header_columns), 1)

    if generic_ratio >= 0.5 or numeric_like_count > 0 or match_count == 0 or textish_columns >= max(len(header_columns) - 1, 1):
        return "data"
    return "header"


def _coerce_sheet_dataframe(data: bytes, sheet_name: str, header_mode: str) -> Tuple[pd.DataFrame, str]:
    mode = (header_mode or "auto").strip().lower()
    if mode not in {"auto", "header", "data"}:
        mode = "auto"

    header_df = pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, header=0, dtype=object)
    data_df = pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, header=None, dtype=object)

    if mode == "header":
        return _clean_dataframe(header_df), "header"
    if mode == "data":
        return _prepare_headerless_dataframe(data_df), "data"

    detected = _detect_sheet_mode(header_df, data_df)
    if detected == "data":
        return _prepare_headerless_dataframe(data_df), "data"
    return _clean_dataframe(header_df), "header"


def read_workbook_metadata(data: bytes) -> List[str]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    return list(wb.sheetnames)


def read_sheet_dataframe(data: bytes, sheet_name: str, header_mode: str = "auto") -> pd.DataFrame:
    df, _ = _coerce_sheet_dataframe(data, sheet_name, header_mode)
    return df


def dataframe_preview(df: pd.DataFrame, rows: int = 5) -> List[Dict[str, Any]]:
    if df.empty:
        return []
    sample = df.head(rows).copy()
    sample = sample.where(pd.notna(sample), None)
    return sample.to_dict(orient="records")


def _score_column_from_data(series: pd.Series) -> Dict[str, float]:
    values = series.dropna()
    values = values[values.astype(str).str.strip() != ""]
    if values.empty:
        return {
            "numeric_ratio": 0.0,
            "text_ratio": 0.0,
            "unique_ratio": 0.0,
            "avg_length": 0.0,
        }
    numeric = to_numeric_series(values)
    numeric_ratio = float(numeric.notna().mean())
    text_values = values[numeric.isna()].astype(str)
    unique_ratio = float(values.astype(str).nunique() / len(values))
    avg_length = float(text_values.str.len().mean()) if not text_values.empty else 0.0
    return {
        "numeric_ratio": numeric_ratio,
        "text_ratio": 1.0 - numeric_ratio,
        "unique_ratio": unique_ratio,
        "avg_length": avg_length,
    }


def detect_default_mapping(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    ranked = detect_field_columns(df.columns)
    mapping: Dict[str, Optional[str]] = {}
    for field, candidates in ranked.items():
        mapping[field] = candidates[0] if candidates else None

    if "investment_name" not in mapping or not mapping.get("investment_name"):
        mapping["investment_name"] = _find_best_text_column(df)

    if "record_date_nav" not in mapping or not mapping.get("record_date_nav"):
        mapping["record_date_nav"] = _find_best_numeric_column(df)

    if mapping.get("sub_asset_class") is None and mapping.get("security_type"):
        mapping["sub_asset_class"] = mapping["security_type"]

    return mapping


def _find_best_text_column(df: pd.DataFrame) -> Optional[str]:
    best_col = None
    best_score = -1.0
    for col in df.columns:
        scores = _score_column_from_data(df[col])
        score = (scores["text_ratio"] * 5.0) + (scores["unique_ratio"] * 3.0) + (scores["avg_length"] / 20.0) - (scores["numeric_ratio"] * 4.0)
        if score > best_score:
            best_score = score
            best_col = col
    if best_score <= 0:
        return None
    return str(best_col)


def _find_best_numeric_column(df: pd.DataFrame) -> Optional[str]:
    best_col = None
    best_score = -1.0
    for col in df.columns:
        scores = _score_column_from_data(df[col])
        score = (scores["numeric_ratio"] * 8.0) + (scores["unique_ratio"] * 0.2) - (scores["avg_length"] / 50.0)
        if score > best_score:
            best_score = score
            best_col = col
    if best_score <= 0.25:
        return None
    return str(best_col)


def infer_sheet_preview(data: bytes, sheet_name: str, header_mode: str = "auto") -> Dict[str, Any]:
    df, detected_mode = _coerce_sheet_dataframe(data, sheet_name, header_mode)
    return {
        "columns": [str(c) for c in df.columns],
        "suggested_columns": detect_default_mapping(df),
        "column_rankings": detect_field_columns(df.columns),
        "sample_rows": dataframe_preview(df),
        "row_count": int(len(df)),
        "header_mode": detected_mode,
    }


def _resolve_column(df: pd.DataFrame, column_name: Optional[str]) -> Optional[str]:
    if not column_name:
        return None
    for col in df.columns:
        if str(col) == str(column_name):
            return col
    return None


def _position_label(row: pd.Series, name_col: Optional[str]) -> str:
    if name_col and name_col in row and pd.notna(row[name_col]):
        text = str(row[name_col]).strip()
        if text:
            return text
    fallback_bits: List[str] = []
    for key in ["asset_class", "security_type", "geography", "sub_asset_class"]:
        val = row.get(key)
        if pd.notna(val) and str(val).strip():
            fallback_bits.append(str(val).strip())
    if fallback_bits:
        return " / ".join(fallback_bits[:3])
    return "Unnamed Position"


def _build_category_breakdown(
    df: pd.DataFrame,
    nav_col: str,
    category_col: Optional[str],
    normalization_rules: Dict[str, str],
) -> List[Dict[str, Any]]:
    if not category_col or category_col not in df.columns:
        return []
    work = df[[nav_col, category_col]].copy()
    work["_category"] = work[category_col].map(lambda v: normalize_category(v, normalization_rules))
    grouped = work.groupby("_category", dropna=False)[nav_col].sum().sort_values(ascending=False)
    total = float(work[nav_col].sum()) or 0.0
    out: List[Dict[str, Any]] = []
    for category, value in grouped.items():
        pct = float(value) / total if total else 0.0
        out.append(
            {
                "label": str(category),
                "value": float(value),
                "percentage": pct,
            }
        )
    return out


def _build_manual_category_breakdown(manual_items: List[Tuple[str, float]]) -> List[Dict[str, Any]]:
    return [
        {
            "label": str(label),
            "value": float(value),
            "percentage": float(value),
        }
        for label, value in manual_items
    ]


def _is_cash_like_row(row: pd.Series) -> bool:
    haystack = " ".join(
        str(value).strip().lower()
        for value in row.tolist()
        if pd.notna(value) and str(value).strip()
    )
    if not haystack:
        return False
    keywords = [
        "cash",
        "cash equivalent",
        "cash equivalents",
        "money market",
        "treasury",
        "t bill",
        "t bills",
        "t-bill",
        "t-bills",
        "mmf",
    ]
    return any(keyword in haystack for keyword in keywords)


def _compute_fund_profile(
    df: pd.DataFrame,
    fund_weight: float,
    normalization_rules: Dict[str, str],
    column_map: Dict[str, Optional[str]],
    manual_families: Dict[str, List[Tuple[str, float]]],
    alias_map: Optional[Dict[str, str]] = None,
    split_set: set = None,
) -> Dict[str, Any]:
    nav_col = _resolve_column(df, column_map.get("record_date_nav"))
    if not nav_col:
        raise ValueError("Missing record date NAV column")

    working = df.copy()
    working["_nav"] = to_numeric_series(working[nav_col]).fillna(0.0)
    working = working[working["_nav"] > 0].copy()
    working["_is_cash"] = working.apply(_is_cash_like_row, axis=1)
    invested = working[~working["_is_cash"]].copy()
    total_nav = float(invested["_nav"].sum())
    if total_nav <= 0:
        raise ValueError("Record date NAV column contains no positive numeric values")

    name_col = _resolve_column(invested, column_map.get("investment_name"))
    aliases = alias_map or {}
    invested["_raw"] = invested.apply(lambda row: _position_label(row, name_col), axis=1)
    invested["_key"] = invested["_raw"].map(lambda r: _effective_key(r, aliases, split_set))
    invested["_project_share"] = (invested["_nav"] / total_nav) * fund_weight

    out: Dict[str, Any] = {
        "total_nav": total_nav,
        "positions": int(invested["_key"].nunique()),
        "position_exposure": [],
        "categories": {},
        "cash_rows": int(working["_is_cash"].sum()),
    }

    # Combine tranches of the same issuer within the fund (e.g. senior + junior debt).
    pos: List[Dict[str, Any]] = []
    for key, sub in invested.groupby("_key", dropna=False):
        value = float(sub["_project_share"].sum())
        variants = sorted({str(x) for x in sub["_raw"].tolist()})
        pos.append({
            "label": _shortest_display(variants) or str(key),
            "value": value,
            "percentage": value,
            "key": str(key),
            "variants": variants,
        })
    pos.sort(key=lambda d: -d["value"])
    out["position_exposure"] = pos

    for family in ["asset_class", "security_type", "geography", "sub_asset_class"]:
        if family in manual_families:
            out["categories"][family] = _build_manual_category_breakdown(manual_families[family])
            continue
        family_col_name = column_map.get(family)
        if family == "sub_asset_class" and not family_col_name:
            family_col_name = column_map.get("security_type")
        family_col = _resolve_column(invested, family_col_name)
        breakdown = _build_category_breakdown(invested, "_nav", family_col, normalization_rules)
        if breakdown:
            out["categories"][family] = breakdown

    return out


def _manual_fund_profile(manual_families: Dict[str, List[Tuple[str, float]]]) -> Dict[str, Any]:
    """Profile for a fund with no workbook: categories come straight from the manual
    overrides (asset class / security type / geography). No positions or NAV."""
    out: Dict[str, Any] = {"total_nav": 0.0, "positions": 0, "position_exposure": [], "categories": {}, "cash_rows": 0}
    for family in ["asset_class", "security_type", "geography", "sub_asset_class"]:
        if family in manual_families:
            out["categories"][family] = _build_manual_category_breakdown(manual_families[family])
    if "sub_asset_class" not in out["categories"] and "security_type" in out["categories"]:
        out["categories"]["sub_asset_class"] = list(out["categories"]["security_type"])
    return out


def compute_project_exposure(
    funds: List[Dict[str, Any]],
    uploads: Dict[str, Dict[str, Any]],
    normalization_rules: Dict[str, str],
    issuer_aliases: Any = "",
    force_splits: Any = "",
) -> Dict[str, Any]:
    if not funds:
        raise ValueError("Add at least one fund before calculating exposures")
    alias_map = parse_issuer_aliases(issuer_aliases)
    split_set = parse_force_splits(force_splits)

    # A fund is usable if it has an imported workbook OR manual exposure overrides.
    # Funds with neither (e.g. just added, not yet imported) are skipped, and weights
    # are normalized across the funds that actually carry composition data.
    prepared: List[tuple] = []
    for fund in funds:
        upload_id = fund.get("upload_id")
        manual_families = parse_manual_overrides(fund.get("manual_category_overrides") or "")
        has_wb = bool(upload_id and upload_id in uploads)
        bid = float(fund.get("bid_amount") or 0.0)
        if bid < 0:
            raise ValueError(f"Bid amount must be non-negative for fund {fund.get('fund_name') or fund.get('filename')}")
        if has_wb or manual_families:
            prepared.append((fund, has_wb, manual_families, bid))
    if not prepared:
        raise ValueError("Add at least one fund with an imported workbook or manual exposures.")
    total_bid = sum(b for _, _, _, b in prepared)
    if total_bid <= 0:
        raise ValueError("Total bid amount must be greater than zero")

    fund_rows: List[Dict[str, Any]] = []
    fund_profiles: List[Dict[str, Any]] = []
    project_categories: Dict[str, Dict[str, float]] = {}
    project_positions: Dict[str, float] = {}          # canonical issuer key -> summed share
    position_variants: Dict[str, Dict[str, int]] = {}  # canonical key -> {raw name: count}

    for fund, has_wb, manual_families, bid in prepared:
        weight = bid / total_bid
        column_map = fund.get("column_map") or {}
        if not column_map.get("sub_asset_class") and column_map.get("security_type"):
            column_map["sub_asset_class"] = column_map.get("security_type")

        if has_wb:
            upload = uploads[fund["upload_id"]]
            sheet_name = fund.get("sheet_name") or upload["default_sheet"]
            header_mode = fund.get("header_mode") or upload.get("header_mode") or "auto"
            df = read_sheet_dataframe(upload["data"], sheet_name, header_mode=header_mode)
            profile = _compute_fund_profile(df, weight, normalization_rules, column_map, manual_families, alias_map, split_set)
            filename = upload["filename"]
        else:
            sheet_name = ""
            header_mode = fund.get("header_mode") or "auto"
            profile = _manual_fund_profile(manual_families)
            filename = fund.get("filename") or ""

        common = {
            "fund_name": fund.get("fund_name") or filename or "Untitled fund",
            "filename": filename,
            "sheet_name": sheet_name,
            "header_mode": header_mode,
            "bid_amount": bid,
            "weight": weight,
            "total_nav": profile["total_nav"],
            "normalized_nav": profile["total_nav"],
            "positions": profile["positions"],
            "cash_rows": profile.get("cash_rows", 0),
            "manual_category_overrides": fund.get("manual_category_overrides") or "",
        }
        fund_profiles.append({**common, "categories": profile.get("categories") or {}, "position_exposure": profile.get("position_exposure") or []})
        fund_rows.append(dict(common))

        for family, breakdown in profile["categories"].items():
            family_bucket = project_categories.setdefault(family, {})
            for item in breakdown:
                family_bucket[item["label"]] = family_bucket.get(item["label"], 0.0) + item["percentage"] * weight

        for item in profile["position_exposure"]:
            key = item.get("key") or _effective_key(item["label"], alias_map, split_set)
            project_positions[key] = project_positions.get(key, 0.0) + item["value"]
            bucket = position_variants.setdefault(key, {})
            for v in (item.get("variants") or [item["label"]]):
                bucket[v] = bucket.get(v, 0) + 1

    category_results: Dict[str, List[Dict[str, Any]]] = {}
    for family, bucket in project_categories.items():
        items = [
            {"label": label, "value": value, "percentage": value}
            for label, value in bucket.items()
        ]
        items.sort(key=lambda item: (-item["value"], item["label"].lower()))
        category_results[family] = items

    top_positions = []
    for key, value in sorted(project_positions.items(), key=lambda kv: (-kv[1], kv[0])):
        variants = sorted(position_variants.get(key, {}).keys())
        top_positions.append({
            "label": _shortest_display(variants) or key,
            "value": value,
            "percentage": value,
            "key": key,
            "variants": variants,
        })

    # Names merged from more than one raw variant, for the review list in the UI.
    position_merges = [
        {"label": item["label"], "variants": item["variants"]}
        for item in top_positions if len(item["variants"]) > 1
    ]

    # Possible same-issuer pairs that did NOT auto-merge: one issuer key is a token-prefix
    # of another (e.g. "noble" vs "noble supply and logistics"). User can confirm via alias.
    keyed = [(it["key"].split(), it["label"]) for it in top_positions if it.get("key")]
    seen = set()
    position_merge_suggestions = []
    for i, (toks_a, label_a) in enumerate(keyed):
        for toks_b, label_b in keyed[i + 1:]:
            shorter, longer = (toks_a, toks_b) if len(toks_a) <= len(toks_b) else (toks_b, toks_a)
            if shorter and len(shorter) < len(longer) and longer[:len(shorter)] == shorter and len(shorter[0]) >= 3:
                pair = tuple(sorted((label_a, label_b)))
                if pair not in seen:
                    seen.add(pair)
                    position_merge_suggestions.append({"a": pair[0], "b": pair[1]})
    position_merge_suggestions = position_merge_suggestions[:20]

    cumulative = 0.0
    top_n = {}
    ordered = [item["value"] for item in top_positions]
    for n in [1, 3, 5, 10]:
        cumulative = sum(ordered[:n])
        top_n[f"top_{n}"] = cumulative
    top_n["remaining"] = max(0.0, 1.0 - sum(ordered[:10]))

    project_sentence = _build_summary_sentence(category_results)

    return {
        "total_bid": total_bid,
        "funds": fund_rows,
        "fund_profiles": fund_profiles,
        "categories": category_results,
        "top_positions": top_positions,
        "top_concentration": top_n,
        "position_merges": position_merges,
        "position_merge_suggestions": position_merge_suggestions,
        "summary_sentence": project_sentence,
    }


def _build_summary_sentence(category_results: Dict[str, List[Dict[str, Any]]]) -> str:
    asset_class = category_results.get("asset_class", [])
    if not asset_class:
        return ""
    top = asset_class[:2]
    parts = [f"{item['label']} ({item['percentage']:.1%})" for item in top]
    if len(parts) == 1:
        return f"The project is primarily exposed to {parts[0]}."
    return f"The project is primarily exposed to {parts[0]} and {parts[1]}."


def top_positions_concentration(top_positions: List[Dict[str, Any]]) -> Dict[str, float]:
    ordered = [float(item["value"]) for item in top_positions]
    return {
        "top_1": sum(ordered[:1]),
        "top_3": sum(ordered[:3]),
        "top_5": sum(ordered[:5]),
        "top_10": sum(ordered[:10]),
        "remaining": max(0.0, 1.0 - sum(ordered[:10])),
    }


def export_json_safe(payload: Dict[str, Any]) -> str:
    return json.dumps(payload, indent=2, sort_keys=True)
