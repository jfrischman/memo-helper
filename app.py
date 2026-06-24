from __future__ import annotations

import cgi
import base64
import io
import json
import os
import threading
import uuid
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Dict, List

from exposure_engine import (
    compute_project_exposure,
    infer_sheet_preview,
    parse_mapping_rules,
    read_workbook_metadata,
)
from memo_export import build_memo_export, update_sections_in_file
from project_store import (
    create_project,
    list_project_summaries,
    load_project,
    read_upload_bytes,
    project_dir,
    save_project,
    store_upload_bytes,
)


HOST = "127.0.0.1"
PORT = int(os.environ.get("PROJECT_BALANCE_PORT", "8799"))

UPLOADS: Dict[str, Dict[str, Any]] = {}


HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Memo Helper</title>
  <style>
    :root {
      --bg: #f4f2ed;
      --panel: #ffffff;
      --panel-2: #f8f6f1;
      --line: #d9d4c9;
      --text: #1f2937;
      --muted: #67707e;
      --accent: #1f5f74;
      --accent-2: #8b5e3c;
      --good: #1f7a4d;
      --bad: #a44444;
      --shadow: 0 10px 30px rgba(31,41,55,.08);
      --radius: 12px;
      --radius-sm: 8px;
      --font: "Segoe UI", Arial, sans-serif;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: var(--font);
      color: var(--text);
      background:
        linear-gradient(180deg, rgba(255,255,255,.75), rgba(255,255,255,.75)),
        radial-gradient(circle at top right, rgba(31,95,116,.08), transparent 32%),
        var(--bg);
    }
    header {
      padding: 24px 28px 12px;
      border-bottom: 1px solid rgba(217,212,201,.8);
      background: rgba(255,255,255,.45);
      backdrop-filter: blur(4px);
      position: sticky;
      top: 0;
      z-index: 2;
    }
    h1 {
      margin: 0 0 6px;
      font-size: 24px;
      line-height: 1.2;
      letter-spacing: 0;
    }
    .sub {
      color: var(--muted);
      font-size: 13px;
      line-height: 1.45;
      max-width: 980px;
    }
    main {
      padding: 20px 28px 40px;
      display: grid;
      gap: 16px;
    }
    .panel {
      background: rgba(255,255,255,.9);
      border: 1px solid var(--line);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      overflow: hidden;
    }
    .panel-head {
      padding: 14px 16px;
      border-bottom: 1px solid var(--line);
      background: linear-gradient(180deg, rgba(248,246,241,1), rgba(255,255,255,1));
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
    }
    .panel-head h2 {
      margin: 0;
      font-size: 15px;
    }
    .panel-body { padding: 16px; }
    .toolbar {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      align-items: center;
    }
    .btn {
      appearance: none;
      border: 1px solid var(--line);
      background: #fff;
      color: var(--text);
      border-radius: 10px;
      padding: 9px 13px;
      font: inherit;
      font-size: 13px;
      cursor: pointer;
      transition: transform .04s ease, border-color .12s ease, box-shadow .12s ease;
    }
    .btn:hover { border-color: #b7b0a3; box-shadow: 0 6px 16px rgba(31,41,55,.06); }
    .btn:active { transform: translateY(1px); }
    .btn.primary {
      background: var(--accent);
      color: white;
      border-color: transparent;
    }
    .btn.secondary {
      background: #f7f3ee;
    }
    .field {
      display: grid;
      gap: 6px;
    }
    label {
      font-size: 12px;
      color: var(--muted);
    }
    input[type="text"], input[type="number"], select, textarea {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 10px;
      background: white;
      padding: 9px 10px;
      font: inherit;
      font-size: 13px;
      color: var(--text);
    }
    textarea {
      min-height: 140px;
      resize: vertical;
      line-height: 1.45;
    }
    input[type="file"] {
      font: inherit;
      font-size: 13px;
    }
    .grid {
      display: grid;
      gap: 14px;
      grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
    }
    .fund-card {
      border: 1px solid var(--line);
      border-radius: var(--radius);
      background: var(--panel);
      overflow: hidden;
    }
    .fund-card .top {
      padding: 12px 14px;
      background: linear-gradient(180deg, var(--panel-2), #fff);
      border-bottom: 1px solid var(--line);
      display: grid;
      gap: 10px;
    }
    .fund-title {
      display: flex;
      justify-content: space-between;
      align-items: start;
      gap: 10px;
    }
    .fund-title strong {
      display: block;
      font-size: 14px;
      margin-bottom: 4px;
    }
    .fund-title .meta {
      color: var(--muted);
      font-size: 12px;
      line-height: 1.35;
      word-break: break-word;
    }
    .fund-body { padding: 14px; display: grid; gap: 12px; }
    .mapping-grid {
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    }
    .preview-wrap {
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 10px;
      background: #fff;
    }
    .manual-editor {
      display: grid;
      gap: 12px;
    }
    .manual-section {
      border: 1px solid var(--line);
      border-radius: 10px;
      background: #fff;
      padding: 12px;
      display: grid;
      gap: 10px;
    }
    .manual-section-head {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 10px;
    }
    .manual-section-head strong {
      font-size: 13px;
    }
    .manual-list {
      display: grid;
      gap: 8px;
    }
    .manual-row {
      display: grid;
      grid-template-columns: minmax(180px, 1.4fr) minmax(90px, 120px) auto auto auto;
      gap: 8px;
      align-items: center;
    }
    .manual-row .manual-label, .manual-row .manual-pct {
      width: 100%;
    }
    .manual-asset-badge {
      font-size: 12px;
      color: var(--accent);
      background: #eef5f8;
      border: 1px solid rgba(31,95,116,.15);
      border-radius: 999px;
      padding: 4px 8px;
      white-space: nowrap;
    }
    .manual-remove {
      padding: 7px 10px;
    }
    table {
      border-collapse: collapse;
      width: 100%;
      font-size: 12px;
    }
    th, td {
      padding: 8px 10px;
      border-bottom: 1px solid #ece7de;
      vertical-align: top;
      text-align: left;
      white-space: nowrap;
    }
    th {
      background: #faf8f4;
      position: sticky;
      top: 0;
      z-index: 1;
    }
    .muted { color: var(--muted); }
    .error {
      color: #7b2626;
      background: #fcf2f2;
      border: 1px solid #edcfcf;
      border-radius: 10px;
      padding: 10px 12px;
      white-space: pre-wrap;
      display: none;
    }
    .stats {
      display: grid;
      gap: 12px;
      grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
    }
    .stat {
      border: 1px solid var(--line);
      border-radius: 12px;
      background: white;
      padding: 14px;
    }
    .stat .k { color: var(--muted); font-size: 12px; margin-bottom: 8px; }
    .stat .v { font-size: 20px; font-weight: 700; }
    .stat .s { color: var(--muted); font-size: 12px; margin-top: 6px; }
    .charts {
      display: grid;
      gap: 14px;
      grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    }
    .chart-card {
      border: 1px solid var(--line);
      border-radius: 12px;
      background: white;
      padding: 14px;
    }
    .chart-card h3 {
      margin: 0 0 10px;
      font-size: 14px;
    }
    .chart-row {
      display: flex;
      gap: 14px;
      align-items: center;
      flex-wrap: wrap;
    }
    .legend {
      display: grid;
      gap: 6px;
      min-width: 180px;
      flex: 1 1 180px;
    }
    .legend-item {
      display: grid;
      grid-template-columns: 14px 1fr auto;
      gap: 8px;
      align-items: center;
      font-size: 12px;
    }
    .swatch {
      width: 14px;
      height: 14px;
      border-radius: 4px;
    }
    .bar-list {
      display: grid;
      gap: 8px;
    }
    .bar-item {
      display: grid;
      grid-template-columns: 180px 1fr 64px;
      gap: 8px;
      align-items: center;
      font-size: 12px;
    }
    .bar-track {
      height: 12px;
      background: #efe9df;
      border-radius: 999px;
      overflow: hidden;
    }
    .bar-fill {
      height: 100%;
      border-radius: 999px;
      background: linear-gradient(90deg, var(--accent), #4d8aa0);
    }
    .section-gap { margin-top: 4px; }
    .small-note { font-size: 12px; color: var(--muted); line-height: 1.45; }
    .download-row { display:flex; gap:10px; flex-wrap:wrap; }
    .drop-zone {
      border: 2px dashed var(--line);
      border-radius: var(--radius-sm);
      padding: 12px;
      transition: border-color .15s, background .15s;
      cursor: default;
    }
    .drop-zone.drag-over {
      border-color: var(--accent);
      background: rgba(31,95,116,.05);
    }
    .drop-zone.drag-over * {
      pointer-events: none;
    }
    .fi-lock-btn {
      background: none; border: none; cursor: pointer;
      font-size: 13px; padding: 0 2px; line-height: 1;
      opacity: 0.25; transition: opacity .15s;
      flex-shrink: 0;
    }
    .fi-lock-btn:hover { opacity: 0.7; }
    .fi-lock-btn.locked { opacity: 1; color: var(--accent); }
    .fi-cell.locked {
      border-color: var(--accent) !important;
      background: rgba(31,95,116,.06);
    }
    .pill {
      display:inline-block;
      padding: 2px 8px;
      border-radius: 999px;
      background: #eef5f8;
      color: var(--accent);
      font-size: 12px;
      border: 1px solid rgba(31,95,116,.15);
    }
    @media (max-width: 760px) {
      header, main { padding-left: 14px; padding-right: 14px; }
      .bar-item { grid-template-columns: 1fr; }
      .bar-track { order: 3; }
    }
  </style>
</head>
<body>
  <header>
    <h1>Memo Helper</h1>
    <div class="sub">
      Start a project, attach it to a memo, upload one workbook per fund, and save the project so the inputs stay with the memo.
    </div>
  </header>
  <main>
    <section class="panel">
      <div class="panel-head">
        <h2>Project</h2>
        <div class="toolbar">
          <button class="btn secondary" id="newProjectBtn">Start new project</button>
          <button class="btn" id="saveProjectBtn">Save project</button>
          <button class="btn primary" id="updateExposuresBtn">Update Exposures</button>
          <button class="btn primary" id="updateFundInfoBtn" style="background:var(--accent-2)">Update Fund Information</button>
          <button class="btn primary" id="updateModelOutputsBtn" style="background:#6b5b95">Update Model Outputs</button>
          <button class="btn" id="updateAllBtn">Update All</button>
          <button class="btn" id="exportMemoBtn">Export memo</button>
        </div>
      </div>
      <div class="panel-body">
        <div class="mapping-grid" style="grid-template-columns:repeat(auto-fit,minmax(220px,1fr));">
          <div class="field">
            <label for="projectSelect">Saved projects</label>
            <select id="projectSelect"></select>
          </div>
          <div class="field">
            <label for="projectNameInput">Project name</label>
            <input type="text" id="projectNameInput" placeholder="Project Balance / Project Blue Torch" />
          </div>
          <div class="field">
            <label for="memoNameInput">Memo name</label>
            <input type="text" id="memoNameInput" placeholder="Project Balance IC Memo" />
          </div>
        </div>
        <div class="mapping-grid section-gap" style="grid-template-columns:repeat(auto-fit,minmax(220px,1fr));">
          <div class="field">
            <label for="memoPathInput">Memo file (SharePoint / OneDrive synced path)</label>
            <div style="display:flex; gap:8px;">
              <input type="text" id="memoPathInput" placeholder="C:\Users\...\OneDrive - GCM Grosvenor\...\Project X IC Memo.docx" style="flex:1;" />
              <button class="btn" id="browseMemoBtn" type="button">Browse…</button>
            </div>
          </div>
          <div class="field">
            <label for="modelXlsxPathInput">Excel model file</label>
            <div style="display:flex; gap:8px;">
              <input type="text" id="modelXlsxPathInput" placeholder="C:\Users\...\Project X - Combined Model.xlsx" style="flex:1;" />
              <button class="btn" id="browseModelBtn" type="button">Browse…</button>
            </div>
          </div>
          <div class="field">
            <label for="projectStatus">Status</label>
            <input type="text" id="projectStatus" readonly value="No project loaded" />
          </div>
        </div>
      </div>
    </section>

    <section class="panel">
      <div class="panel-head">
        <h2>Funds</h2>
        <div class="toolbar">
          <button class="btn primary" id="addFundBtn">Add fund</button>
          <button class="btn" id="clearBtn">Clear</button>
        </div>
      </div>
      <div class="panel-body">
        <div class="small-note">
          Add a fund, then (optionally) import its holdings workbook within the fund card. A fund without a
          workbook can still contribute via its manual category mix. Imported sheets are expected to have a header row.
        </div>
        <div id="errorBox" class="error section-gap"></div>
        <div id="fundList" class="grid section-gap"></div>
      </div>
    </section>

    <section class="panel">
      <div class="panel-head">
        <h2>Fund Information</h2>
        <div class="toolbar">
          <button class="btn secondary" id="parseProjFilesBtn">Parse project files</button>
        </div>
      </div>
      <div class="panel-body">
        <div class="grid" style="grid-template-columns:repeat(auto-fit,minmax(260px,1fr));">
          <div class="field">
            <label for="openaiKeyInput">OpenAI API Key</label>
            <input type="password" id="openaiKeyInput" placeholder="sk-proj-..." />
          </div>
          <div class="field">
            <label>Project files (LP NAV / Unfunded / Commits source)</label>
            <div id="projDropZone" class="drop-zone" style="display:flex;flex-direction:column;gap:8px;">
              <div style="display:flex;gap:8px;align-items:center;">
                <label class="btn secondary" for="projFileUpload" style="display:inline-flex;align-items:center;gap:6px;white-space:nowrap;"><span>Choose files</span><input type="file" id="projFileUpload" accept=".pdf,.xlsx,.docx" multiple style="display:none"/></label>
                <span class="small-note" style="flex:1">or drag &amp; drop PDFs here</span>
              </div>
              <input type="text" id="projFilesPathInput" placeholder="Or paste folder / file path…" style="width:100%"/>
              <div id="projFilesList" class="small-note"></div>
            </div>
          </div>
        </div>
        <div class="section-gap" id="fundInfoParseStatus" style="display:none;align-items:center;gap:10px;padding:10px 0;">
          <svg width="22" height="22" viewBox="0 0 22 22" style="flex-shrink:0;animation:fi-spin 1s linear infinite">
            <circle cx="11" cy="11" r="9" fill="none" stroke="#d9d4c9" stroke-width="3"/>
            <path d="M11 2 A9 9 0 0 1 20 11" fill="none" stroke="var(--accent)" stroke-width="3" stroke-linecap="round"/>
          </svg>
          <span id="fundInfoParseMsg" style="font-size:13px;color:var(--muted)">Parsing…</span>
        </div>
        <style>@keyframes fi-spin{to{transform:rotate(360deg)}}</style>
        <div id="fundInfoErrorBox" class="error section-gap" style="display:none;"></div>
        <div class="section-gap" id="fundInfoTableWrap">
          <div class="small-note">Add funds and parse documents to populate the fund information table.</div>
        </div>
        <div class="section-gap" id="fundDocsList"></div>
      </div>
    </section>

    <section class="panel">
      <div class="panel-head">
        <h2>Normalization Rules</h2>
        <div class="toolbar">
          <button class="btn primary" id="calcBtn">Calculate exposures</button>
        </div>
      </div>
      <div class="panel-body">
        <div class="field">
          <label for="rulesBox">Category rules — one per line, <code>raw => normalized</code>.</label>
          <textarea id="rulesBox" placeholder="Corp Lending => Corporate Lending&#10;Corp. Lending => Corporate Lending&#10;Special Sits => Special Situations"></textarea>
        </div>
        <div class="field section-gap">
          <label for="aliasesBox">Issuer aliases — force names to one issuer, <code>variant => issuer</code>. The same issuer named differently across funds is auto-combined; use this only for cases it can't catch (e.g. abbreviations).</label>
          <textarea id="aliasesBox" placeholder="Noble => Noble Supply and Logistics&#10;ABC Co; ABC Holdings => ABC"></textarea>
        </div>
        <div class="field section-gap">
          <label for="labelColorsBox">Pie label colors — one line per category, <code>Category = white</code> or <code>Category = black</code>. Auto-filled from results (≥7% = white, &lt;7% = black). Edit to override individual slices.</label>
          <textarea id="labelColorsBox" placeholder="Corporate Lending = white&#10;ABS = black&#10;Special Situations = black&#10;North America = white&#10;Europe = black"></textarea>
        </div>
        <div class="field section-gap">
          <label for="splitNamesBox">Split names — one raw name per line. These will NOT be merged with other issuers even if they share the same root name. Add via the "Unsplit" button in results, or type directly.</label>
          <textarea id="splitNamesBox" placeholder="CREO Group Inc.&#10;CREO Group (fka Nursery Supplies)"></textarea>
        </div>
      </div>
    </section>

    <section class="panel">
      <div class="panel-head">
        <h2>Exposure</h2>
        <div class="toolbar">
          <label class="btn secondary" style="display:inline-flex;align-items:center;gap:8px;">
            <input id="liveUpdateToggle" type="checkbox" checked />
            <span>Live chart updates</span>
          </label>
          <div class="download-row">
          <button class="btn" id="downloadJsonBtn">Download JSON</button>
          <button class="btn" id="downloadExcelBtn">Download positions Excel</button>
          </div>
        </div>
      </div>
      <div class="panel-body" id="resultsPane">
        <div class="small-note">No exposure data yet. Add a fund and click calculate.</div>
      </div>
    </section>
  </main>

  <template id="fundCardTemplate">
    <div class="fund-card">
      <div class="top">
        <div class="fund-title">
          <div>
            <strong class="fund-name">Fund</strong>
            <div class="meta fund-meta"></div>
          </div>
          <div style="display:flex; gap:8px; align-items:center;">
            <span class="pill fund-status">no workbook</span>
            <button class="btn fund-remove" type="button">Remove</button>
          </div>
        </div>
        <div class="toolbar" style="gap:8px;">
          <label class="btn secondary fund-import-label" style="display:inline-flex;align-items:center;gap:8px;">
            <span class="fund-import-text">Import Excel</span>
            <input type="file" class="fund-import" accept=".xlsx,.xlsm" style="display:none" />
          </label>
        </div>
        <div class="mapping-grid">
          <div class="field"><label>Display label</label><input type="text" class="fund-label" /></div>
          <div class="field"><label>Abbreviation (for memo labels, e.g. "BT III")</label><input type="text" class="fund-abbrev" placeholder="e.g. BT III" /></div>
          <div class="field"><label>Sheet</label><select class="fund-sheet"></select></div>
          <div class="field"><label>Header mode</label><select class="fund-header-mode"></select></div>
          <div class="field"><label>Bid amount</label><input type="number" min="0" step="any" class="fund-bid" /></div>
        </div>
      </div>
      <div class="fund-body">
        <div class="field">
          <label>Manual category overrides</label>
          <div class="manual-editor"></div>
          <div class="small-note">Use the dropdowns to set geography, asset class, and security type mixes. Security type inherits the asset class color automatically.</div>
        </div>
        <div class="mapping-grid fund-mapping"></div>
        <div class="preview-wrap"><table class="preview-table"></table></div>
      </div>
    </div>
  </template>

  <script>
    window.addEventListener('error', (event) => {
      document.body.dataset.scriptError = event.message || 'unknown error';
    });
    window.addEventListener('unhandledrejection', (event) => {
      document.body.dataset.scriptError = event.reason ? String(event.reason) : 'unhandled rejection';
    });

    const FIELD_LABELS = {
      investment_name: 'Investment name',
      record_date_nav: 'Record date NAV',
      asset_class: 'Asset class',
      security_type: 'Security type',
      geography: 'Geography',
    };
    const GEOGRAPHY_OPTIONS = ['North America', 'Europe', 'Other'];
    const ASSET_CLASS_OPTIONS = ['Corporate Lending', 'ABS', 'Special Situations'];
    const SECURITY_TYPE_OPTIONS = [
      { label: 'Direct Lending', asset_class: 'Corporate Lending' },
      { label: 'Other Senior Lending', asset_class: 'Corporate Lending' },
      { label: 'Opportunistic / Junior', asset_class: 'Corporate Lending' },
      { label: 'Distressed', asset_class: 'Corporate Lending' },
      { label: 'Corporate Equity', asset_class: 'Corporate Lending' },
      { label: 'CLOs', asset_class: 'ABS' },
      { label: 'Regulatory Capital', asset_class: 'ABS' },
      { label: 'Commercial RE (Debt)', asset_class: 'ABS' },
      { label: 'Residential RE', asset_class: 'ABS' },
      { label: 'Consumer', asset_class: 'ABS' },
      { label: 'Hard Assets', asset_class: 'ABS' },
      { label: 'Specialty Lending', asset_class: 'ABS' },
      { label: 'Commercial RE (Equity)', asset_class: 'Special Situations' },
      { label: 'Commercial RE (Non-Perf)', asset_class: 'Special Situations' },
      { label: 'Equity', asset_class: 'Special Situations' },
    ];
    const ASSET_CLASS_COLORS = {
      'Corporate Lending': '#1f5f74',
      ABS: '#8b5e3c',
      'Special Situations': '#5d7d4e',
    };
    const GEOGRAPHY_COLORS = {
      'North America': '#4f6fb5',
      Europe: '#c27c3d',
      Other: '#8a4f69',
    };
    const SECURITY_TYPE_TO_ASSET_CLASS = Object.fromEntries(
      SECURITY_TYPE_OPTIONS.map((item) => [item.label, item.asset_class])
    );
    const appState = {
      projects: [],
      project: null,
      funds: [],
      result: null,
      saveTimer: null,
      chartTimer: null,
      calculating: false,
      loadingProject: false,
    };

    const projectSelect = document.getElementById('projectSelect');
    const newProjectBtn = document.getElementById('newProjectBtn');
    const saveProjectBtn = document.getElementById('saveProjectBtn');
    const exportMemoBtn = document.getElementById('exportMemoBtn');
    const projectNameInput = document.getElementById('projectNameInput');
    const memoNameInput = document.getElementById('memoNameInput');
    const memoPathInput = document.getElementById('memoPathInput');
    const browseMemoBtn = document.getElementById('browseMemoBtn');
    const modelXlsxPathInput = document.getElementById('modelXlsxPathInput');
    const browseModelBtn = document.getElementById('browseModelBtn');
    const updateExposuresBtn = document.getElementById('updateExposuresBtn');
    const updateModelOutputsBtn = document.getElementById('updateModelOutputsBtn');
    const updateFundInfoBtn = document.getElementById('updateFundInfoBtn');
    const updateAllBtn = document.getElementById('updateAllBtn');
    const openaiKeyInput = document.getElementById('openaiKeyInput');
    const projFileUpload = document.getElementById('projFileUpload');
    const projFilesPathInput = document.getElementById('projFilesPathInput');
    const projFilesList = document.getElementById('projFilesList');
    const fundInfoTableWrap = document.getElementById('fundInfoTableWrap');
    const fundDocsList = document.getElementById('fundDocsList');
    const parseProjFilesBtn = document.getElementById('parseProjFilesBtn');
    const fundInfoParseStatus = document.getElementById('fundInfoParseStatus');
    const fundInfoParseMsg = document.getElementById('fundInfoParseMsg');

    // Merge parsed fields into a fund_info object, never overwriting existing
    // non-null values with null, and never touching the override.
    function mergeFields(fi, newFields) {
      if (!fi.fields) fi.fields = {};
      Object.entries(newFields).forEach(([k, v]) => {
        if (!fi.fields[k]) fi.fields[k] = {value: null, source: null, override: null, locked: false};
        if (fi.fields[k].locked) return;  // protected — skip entirely
        if (v && v.value != null) {
          fi.fields[k].value  = v.value;
          fi.fields[k].source = v.source || fi.fields[k].source;
        }
        // Always carry through extra keys (perpetuity, perpetuity_note)
        Object.keys(v || {}).forEach(prop => {
          if (prop !== 'value' && prop !== 'source' && prop !== 'override') {
            fi.fields[k][prop] = v[prop];
          }
        });
      });
    }

    function setupDropZone(el, onFiles) {
      el.addEventListener('dragenter', (e) => { e.preventDefault(); el.classList.add('drag-over'); });
      el.addEventListener('dragover',  (e) => { e.preventDefault(); if (e.dataTransfer) e.dataTransfer.dropEffect = 'copy'; el.classList.add('drag-over'); });
      el.addEventListener('dragleave', (e) => { if (!el.contains(e.relatedTarget)) el.classList.remove('drag-over'); });
      el.addEventListener('drop', (e) => {
        e.preventDefault(); el.classList.remove('drag-over');
        const rawFiles = e.dataTransfer && e.dataTransfer.files ? Array.from(e.dataTransfer.files) : [];
        const files = rawFiles.filter(f => /\.(pdf|xlsx|docx)$/i.test(f.name));
        if (files.length) onFiles(files); else showFundInfoError('Drop PDF, XLSX, or DOCX files only.');
      });
    }

    // Project files drop zone
    const projDropZone = document.getElementById('projDropZone');
    if (projDropZone) {
      setupDropZone(projDropZone, async (files) => {
        if (!appState.project) return;
        const form = new FormData();
        form.append('project_id', appState.project.project_id || '');
        files.forEach(f => form.append('files', f));
        const r = await fetch('/api/fund_info/upload_proj_files', {method:'POST', body:form});
        const d = await r.json();
        if (!r.ok) { showError(d.error || 'Upload failed'); return; }
        if (!appState.project.project_file_paths) appState.project.project_file_paths = [];
        (d.paths || []).forEach(p => { if (!appState.project.project_file_paths.includes(p)) appState.project.project_file_paths.push(p); });
        projFilesPathInput.value = appState.project.project_file_paths.join('\n');
        projFilesList.textContent = `${appState.project.project_file_paths.length} file(s) ready.`;
        scheduleProjectSave();
      });
    }

    function wireDocDropZones() {
      document.querySelectorAll('.fi-drop-zone').forEach((zone) => {
        if (zone._dropWired) return;
        zone._dropWired = true;
        const fname = zone.dataset.fund;
        const dtype = zone.dataset.dtype;
        setupDropZone(zone, (files) => uploadFundDoc(fname, dtype, files[0]));
      });
    }

    const fundInfoErrorBox = document.getElementById('fundInfoErrorBox');

    function showFundInfoError(msg) {
      if (!msg) { fundInfoErrorBox.style.display = 'none'; fundInfoErrorBox.textContent = ''; return; }
      fundInfoErrorBox.textContent = msg;
      fundInfoErrorBox.style.display = 'block';
      fundInfoErrorBox.scrollIntoView({behavior:'smooth', block:'nearest'});
    }

    function showParseSpinner(msg) {
      showFundInfoError('');
      fundInfoParseMsg.textContent = msg || 'Parsing…';
      fundInfoParseStatus.style.display = 'flex';
    }
    function hideParseSpinner(msg) {
      fundInfoParseStatus.style.display = 'none';
      if (msg) setProjectStatus(msg);
    }
    const projectStatus = document.getElementById('projectStatus');
    const liveUpdateToggle = document.getElementById('liveUpdateToggle');
    const addFundBtn = document.getElementById('addFundBtn');
    const clearBtn = document.getElementById('clearBtn');
    const calcBtn = document.getElementById('calcBtn');
    const fundList = document.getElementById('fundList');
    const errorBox = document.getElementById('errorBox');
    const rulesBox = document.getElementById('rulesBox');
    const aliasesBox = document.getElementById('aliasesBox');
    const labelColorsBox = document.getElementById('labelColorsBox');
    const splitNamesBox = document.getElementById('splitNamesBox');
    const resultsPane = document.getElementById('resultsPane');
    const downloadJsonBtn = document.getElementById('downloadJsonBtn');
    const downloadExcelBtn = document.getElementById('downloadExcelBtn');

    function showError(message) {
      if (!message) {
        errorBox.style.display = 'none';
        errorBox.textContent = '';
        return;
      }
      errorBox.style.display = 'block';
      errorBox.textContent = message;
    }

    function escapeHtml(value) {
      return String(value)
        .replaceAll('&', '&amp;')
        .replaceAll('<', '&lt;')
        .replaceAll('>', '&gt;')
        .replaceAll('"', '&quot;')
        .replaceAll("'", '&#39;');
    }

    function shortId(value) {
      return String(value).slice(0, 8);
    }

    function defaultLabel(filename) {
      return filename.replace(/\.[^.]+$/, '');
    }

    function setProjectStatus(message) {
      projectStatus.value = message || '';
    }

    function activeProjectId() {
      return appState.project ? appState.project.project_id : '';
    }

    function projectPayloadFromState() {
      const funds = Array.from(document.querySelectorAll('.fund-card')).map(collectFundState);
      return {
        project_id: appState.project ? appState.project.project_id : null,
        project_name: projectNameInput.value.trim() || 'Untitled project',
        memo_name: memoNameInput.value.trim() || '',
        memo_file_path: memoPathInput.value.trim(),
        model_xlsx_path: modelXlsxPathInput.value.trim(),
        rules: rulesBox.value || '',
        issuer_aliases: aliasesBox.value || '',
        label_colors: labelColorsBox.value || '',
        force_splits: splitNamesBox.value || '',
        openai_key: openaiKeyInput.value.trim(),
        project_file_paths: projFilesPathInput.value.trim().split('\n').map(s=>s.trim()).filter(Boolean),
        fund_infos: getFundInfos(),
        funds,
      };
    }

    function hasPositiveBid() {
      return Array.from(document.querySelectorAll('.fund-card')).some((card) => {
        const bid = Number(card.querySelector('.fund-bid')?.value || 0);
        return isFinite(bid) && bid > 0;
      });
    }

    function populateProjectSelect(projects, selectedId) {
      projectSelect.innerHTML = '';
      const empty = document.createElement('option');
      empty.value = '';
      empty.textContent = '(new project)';
      projectSelect.appendChild(empty);
      projects.forEach((project) => {
        const option = document.createElement('option');
        option.value = project.project_id;
        option.textContent = `${project.project_name || 'Untitled project'}${project.memo_name ? ` · ${project.memo_name}` : ''}`;
        if (project.project_id === selectedId) option.selected = true;
        projectSelect.appendChild(option);
      });
    }

    async function loadProjectList(selectProjectId) {
      const resp = await fetch('/api/projects');
      const data = await resp.json();
      if (!resp.ok) {
        throw new Error(data.error || 'Unable to load projects.');
      }
      appState.projects = data.projects || [];
      populateProjectSelect(appState.projects, selectProjectId || (appState.project && appState.project.project_id));
    }

    function fillProjectForm(project) {
      appState.project = project;
      projectNameInput.value = project.project_name || 'Untitled project';
      memoNameInput.value = project.memo_name || '';
      memoPathInput.value = project.memo_file_path || '';
      modelXlsxPathInput.value = project.model_xlsx_path || '';
      rulesBox.value = project.rules || '';
      aliasesBox.value = project.issuer_aliases || '';
      labelColorsBox.value = project.label_colors || '';
      splitNamesBox.value = project.force_splits || '';
      openaiKeyInput.value = project.openai_key || '';
      projFilesPathInput.value = (project.project_file_paths || []).join('\n');
      setProjectStatus(project.project_id ? `Project saved: ${project.project_name || 'Untitled project'}` : 'No project loaded');
      populateProjectSelect(appState.projects, project.project_id);
    }

    async function startNewProject() {
      showError('');
      const resp = await fetch('/api/projects/new', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({
          project_name: projectNameInput.value.trim() || 'Untitled project',
          memo_name: memoNameInput.value.trim() || '',
        }),
      });
      const data = await resp.json();
      if (!resp.ok) {
        throw new Error(data.error || 'Unable to create project.');
      }
      appState.funds = [];
      appState.result = null;
      fillProjectForm(data.project);
      memoPathInput.value = '';
      modelXlsxPathInput.value = '';
      fundList.innerHTML = '';
      resultsPane.innerHTML = '<div class="small-note">New project started. Add a fund to begin.</div>';
      await loadProjectList(data.project.project_id);
    }

    async function openProject(projectId) {
      if (!projectId) {
        appState.project = null;
        appState.funds = [];
        appState.result = null;
        projectNameInput.value = '';
        memoNameInput.value = '';
        memoPathInput.value = '';
        modelXlsxPathInput.value = '';
        rulesBox.value = '';
        aliasesBox.value = '';
        labelColorsBox.value = '';
        splitNamesBox.value = '';
        openaiKeyInput.value = '';
        projFilesPathInput.value = '';
        fundInfoTableWrap.innerHTML = '<div class="small-note">Add funds and parse documents to populate the fund information table.</div>';
        fundDocsList.innerHTML = '';
        fundList.innerHTML = '';
        resultsPane.innerHTML = '<div class="small-note">No exposure data yet. Add a fund and click calculate.</div>';
        setProjectStatus('No project loaded');
        populateProjectSelect(appState.projects, '');
        return;
      }
      appState.loadingProject = true;
      try {
        const resp = await fetch('/api/projects/open', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({ project_id: projectId }),
        });
        const data = await resp.json();
        if (!resp.ok) {
          throw new Error(data.error || 'Unable to open project.');
        }
        appState.funds = data.funds || [];
        appState.result = null;
        fillProjectForm(data.project || {});
        renderFunds();
        resultsPane.innerHTML = '<div class="small-note">Project opened. Load or adjust workbooks, then calculate exposures.</div>';
        await loadProjectList(projectId);
      } finally {
        appState.loadingProject = false;
      }
    }

    function scheduleProjectSave() {
      if (appState.loadingProject) return;
      clearTimeout(appState.saveTimer);
      appState.saveTimer = setTimeout(() => {
        saveProject().catch((err) => showError(err.message || String(err)));
      }, 500);
    }

    function scheduleChartRefresh() {
      if (!liveUpdateToggle || !liveUpdateToggle.checked || appState.loadingProject) return;
      clearTimeout(appState.chartTimer);
      appState.chartTimer = setTimeout(() => {
        if (!liveUpdateToggle.checked || appState.loadingProject || appState.calculating) return;
        if (!hasPositiveBid()) return;
        calculate().catch((err) => showError(err.message || String(err)));
      }, 450);
    }

    async function saveProject() {
      if (!appState.project) return;
      const payload = projectPayloadFromState();
      const resp = await fetch('/api/projects/save', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(payload),
      });
      const data = await resp.json();
      if (!resp.ok) {
        throw new Error(data.error || 'Unable to save project.');
      }
      appState.project = data.project;
      setProjectStatus(`Project saved: ${data.project.project_name || 'Untitled project'}`);
      populateProjectSelect(appState.projects, data.project.project_id);
    }

    async function exportMemo() {
      if (!appState.project) {
        showError('Start or open a project first.');
        return;
      }
      const payload = projectPayloadFromState();
      const resp = await fetch('/api/projects/export_memo', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(payload),
      });
      if (!resp.ok) {
        const data = await resp.json().catch(() => ({}));
        throw new Error(data.error || 'Unable to export memo.');
      }
      const blob = await resp.blob();
      const cd = resp.headers.get('Content-Disposition') || '';
      const match = cd.match(/filename="?([^";]+)"?/i);
      const filename = match ? match[1] : 'memo_helper_export.docx';
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url;
      a.download = filename;
      a.click();
      URL.revokeObjectURL(url);
    }

    function normalizeManualState(raw) {
      const empty = { geography: [], asset_class: [], security_type: [] };
      if (!raw) return empty;
      if (typeof raw === 'object' && !Array.isArray(raw)) {
        return {
          geography: normalizeManualRows(raw.geography || []),
          asset_class: normalizeManualRows(raw.asset_class || []),
          security_type: normalizeManualRows(raw.security_type || []),
        };
      }
      if (typeof raw === 'string') {
        const parsed = parseLegacyManualText(raw);
        return {
          geography: normalizeManualRows(parsed.geography || []),
          asset_class: normalizeManualRows(parsed.asset_class || []),
          security_type: normalizeManualRows(parsed.security_type || []),
        };
      }
      return empty;
    }

    function parseLegacyManualText(text) {
      const out = {};
      const lines = String(text || '').split(/\r?\n/);
      lines.forEach((rawLine) => {
        const line = rawLine.trim();
        if (!line) return;
        const match = line.match(/^(geography|asset\s*class|asset_class|security\s*type|security_type)\s*[:=]\s*(.+)$/i);
        const family = match ? match[1] : 'geography';
        const payload = match ? match[2] : line;
        const rows = payload.split(/[;,]/).map((part) => part.trim()).filter(Boolean).map((part) => {
          const pieces = part.split(/\s*(?:=>|->|=)\s*/);
          let label = '';
          let pct = '';
          if (pieces.length >= 2) {
            label = pieces[0];
            pct = pieces.slice(1).join('=');
          } else {
            const tail = part.match(/^(.*)\s+([0-9.]+%?)$/);
            if (tail) {
              label = tail[1];
              pct = tail[2];
            }
          }
          return { label: label.trim(), pct: parseFloat(String(pct).replace('%', '')) || 0 };
        }).filter((item) => item.label);
        const familyKey = family.replace(/\s+/g, '_').toLowerCase();
        out[familyKey] = rows;
      });
      return out;
    }

    function normalizeManualRows(rows) {
      return (rows || []).map((row) => ({
        label: row.label || row.name || row.value || '',
        pct: displayPctValue(row.pct ?? row.percentage ?? row.value ?? 0),
      })).filter((row) => row.label);
    }

    function rowValueToPct(value) {
      const num = Number(value || 0);
      return isFinite(num) ? num : 0;
    }

    function displayPctValue(value) {
      const num = Number(value || 0);
      if (!isFinite(num)) return 0;
      return num > 0 && num <= 1 ? num * 100 : num;
    }

    function assetClassForSecurityType(label) {
      return SECURITY_TYPE_TO_ASSET_CLASS[label] || '';
    }

    function createOptionList(options, selectedValue) {
      return options.map((optionValue) => {
        const selected = optionValue === selectedValue ? ' selected' : '';
        return `<option value="${escapeHtml(optionValue)}"${selected}>${escapeHtml(optionValue)}</option>`;
      }).join('');
    }

    function createSecurityTypeOptionGroups(selectedValue) {
      const groups = {
        'Corporate Lending': SECURITY_TYPE_OPTIONS.filter((item) => item.asset_class === 'Corporate Lending'),
        ABS: SECURITY_TYPE_OPTIONS.filter((item) => item.asset_class === 'ABS'),
        'Special Situations': SECURITY_TYPE_OPTIONS.filter((item) => item.asset_class === 'Special Situations'),
      };
      return Object.entries(groups).map(([groupLabel, items]) => {
        const options = items.map((item) => {
          const selected = item.label === selectedValue ? ' selected' : '';
          return `<option value="${escapeHtml(item.label)}"${selected}>${escapeHtml(item.label)}</option>`;
        }).join('');
        return `<optgroup label="${escapeHtml(groupLabel)}">${options}</optgroup>`;
      }).join('');
    }

    function makeManualRow(fund, family, row) {
      const wrapper = document.createElement('div');
      wrapper.className = 'manual-row';
      wrapper.dataset.family = family;
      const labelSelect = document.createElement('select');
      labelSelect.className = 'manual-label';
      if (family === 'geography') {
        labelSelect.innerHTML = `<option value="">Select geography</option>${createOptionList(GEOGRAPHY_OPTIONS, row.label || '')}`;
      } else if (family === 'asset_class') {
        labelSelect.innerHTML = `<option value="">Select asset class</option>${createOptionList(ASSET_CLASS_OPTIONS, row.label || '')}`;
      } else {
        labelSelect.innerHTML = `<option value="">Select security type</option>${createSecurityTypeOptionGroups(row.label || '')}`;
      }
      const pctInput = document.createElement('input');
      pctInput.type = 'number';
      pctInput.min = '0';
      pctInput.step = 'any';
      pctInput.className = 'manual-pct';
      pctInput.value = row.pct !== undefined && row.pct !== null && row.pct !== '' ? displayPctValue(row.pct) : '';
      const pctLabel = document.createElement('span');
      pctLabel.className = 'manual-pct-label';
      pctLabel.textContent = '%';
      const removeBtn = document.createElement('button');
      removeBtn.type = 'button';
      removeBtn.className = 'btn manual-remove';
      removeBtn.textContent = 'Remove';
      const assetBadge = document.createElement('span');
      assetBadge.className = 'manual-asset-badge';
      if (family === 'security_type') {
        assetBadge.textContent = assetClassForSecurityType(row.label || '') || 'Asset class';
      } else {
        assetBadge.textContent = '';
      }
      labelSelect.addEventListener('change', () => {
        if (family === 'security_type') {
          assetBadge.textContent = assetClassForSecurityType(labelSelect.value || '') || 'Asset class';
        }
        syncManualState(fund, wrapper.closest('.manual-editor'));
        scheduleProjectSave();
      });
      pctInput.addEventListener('input', () => {
        syncManualState(fund, wrapper.closest('.manual-editor'));
        scheduleProjectSave();
      });
      removeBtn.addEventListener('click', () => {
        wrapper.remove();
        syncManualState(fund, wrapper.closest('.manual-editor'));
        scheduleProjectSave();
      });
      wrapper.appendChild(labelSelect);
      wrapper.appendChild(pctInput);
      wrapper.appendChild(pctLabel);
      if (family === 'security_type') {
        wrapper.appendChild(assetBadge);
      }
      wrapper.appendChild(removeBtn);
      return wrapper;
    }

    function buildManualEditor(fund, container) {
      const state = normalizeManualState(fund.manual_category_overrides);
      fund.manual_category_overrides = state;
      container.innerHTML = '';
      const sections = [
        { key: 'geography', title: 'Geography', options: GEOGRAPHY_OPTIONS },
        { key: 'asset_class', title: 'Asset class', options: ASSET_CLASS_OPTIONS },
        { key: 'security_type', title: 'Security type', options: SECURITY_TYPE_OPTIONS.map((item) => item.label) },
      ];
      sections.forEach((section) => {
        const block = document.createElement('div');
        block.className = 'manual-section';
        block.dataset.family = section.key;
        const head = document.createElement('div');
        head.className = 'manual-section-head';
        head.innerHTML = `<strong>${escapeHtml(section.title)}</strong>`;
        const addBtn = document.createElement('button');
        addBtn.type = 'button';
        addBtn.className = 'btn secondary';
        addBtn.textContent = 'Add row';
        const list = document.createElement('div');
        list.className = 'manual-list';
        const initialRows = (state[section.key] || []).length ? (state[section.key] || []) : [{ label: '', pct: '' }];
        initialRows.forEach((row) => list.appendChild(makeManualRow(fund, section.key, row)));
        addBtn.addEventListener('click', () => {
          list.appendChild(makeManualRow(fund, section.key, { label: '', pct: '' }));
          syncManualState(fund, container);
          scheduleProjectSave();
        });
        head.appendChild(addBtn);
        block.appendChild(head);
        block.appendChild(list);
        container.appendChild(block);
      });
    }

    function syncManualState(fund, container) {
      if (!container) return;
      const out = { geography: [], asset_class: [], security_type: [] };
      container.querySelectorAll('.manual-section').forEach((section) => {
        const list = section.querySelector('.manual-list');
        if (!list) return;
        const family = section.dataset.family || 'geography';
        list.querySelectorAll('.manual-row').forEach((row) => {
          const label = row.querySelector('.manual-label')?.value || '';
          const pct = row.querySelector('.manual-pct')?.value || '';
          if (!label) return;
          const parsed = Number(pct);
          out[family].push({
            label,
            pct: Number.isFinite(parsed) ? parsed / 100.0 : 0,
          });
        });
      });
      fund.manual_category_overrides = out;
    }

    function populateHeaderModeSelect(select, selectedValue) {
      const modes = [
        ['auto', 'Auto-detect'],
        ['header', 'First row has headers'],
        ['data', 'First row is data'],
      ];
      select.innerHTML = '';
      modes.forEach(([value, label]) => {
        const option = document.createElement('option');
        option.value = value;
        option.textContent = label;
        if (value === selectedValue) option.selected = true;
        select.appendChild(option);
      });
    }

    function addFund() {
      const fund = {
        upload_id: 'tmp-' + Math.random().toString(36).slice(2, 10),
        fund_name: 'Fund ' + (appState.funds.length + 1),
        bid_amount: 0,
        manual_category_overrides: { geography: [], asset_class: [], security_type: [] },
        mapping: {},
        preview: null,
        sheet_names: [],
        default_sheet: '',
        header_mode: 'auto',
      };
      appState.funds.push(fund);
      renderFunds();
      scheduleProjectSave();
    }

    function removeFund(fund) {
      appState.funds = appState.funds.filter((f) => f !== fund);
      renderFunds();
      scheduleProjectSave();
      scheduleChartRefresh();
    }

    async function importFundExcel(fund, file) {
      showError('');
      const form = new FormData();
      if (activeProjectId()) form.append('project_id', activeProjectId());
      form.append('files', file);
      const resp = await fetch('/api/load', { method: 'POST', body: form });
      const data = await resp.json();
      if (!resp.ok) throw new Error(data.error || 'Unable to import workbook.');
      const loaded = (data.funds && data.funds[0]) || null;
      if (!loaded) throw new Error('No workbook loaded.');
      // merge the imported workbook into this fund (keep its name/bid/manual mix)
      fund.upload_id = loaded.upload_id;
      fund.filename = loaded.filename;
      fund.sheet_names = loaded.sheet_names;
      fund.default_sheet = loaded.default_sheet;
      fund.sheet_name = loaded.default_sheet;
      fund.header_mode = loaded.header_mode;
      fund.row_count = loaded.row_count;
      fund.preview = loaded.preview;
      fund.mapping = loaded.mapping;
      renderFunds();
      appState.result = null;
      scheduleProjectSave();
      scheduleChartRefresh();
    }

    async function updateExposures() {
      if (!appState.project) { showError('Start or open a project first.'); return; }
      if (!memoPathInput.value.trim()) { showError('Point to a memo file first (Browse next to "Memo file").'); return; }
      const payload = projectPayloadFromState();
      updateExposuresBtn.disabled = true;
      setProjectStatus('Updating exposures in memo...');
      try {
        const resp = await fetch('/api/projects/update_exposures', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify(payload),
        });
        const data = await resp.json();
        if (!resp.ok) throw new Error(data.error || 'Unable to update exposures.');
        setProjectStatus('Exposures updated in: ' + (data.memo_file_path || 'memo'));
      } finally {
        updateExposuresBtn.disabled = false;
      }
    }

    async function browseMemo() {
      const resp = await fetch('/api/pick_file', { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: '{}' });
      const data = await resp.json();
      if (data.error) showError(data.error);
      if (data.path) {
        memoPathInput.value = data.path;
        if (appState.project) { setProjectStatus('Unsaved changes'); scheduleProjectSave(); }
      }
    }

    async function browseModel() {
      const resp = await fetch('/api/pick_file', { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify({filter: 'excel'}) });
      const data = await resp.json();
      if (data.error) showError(data.error);
      if (data.path) {
        modelXlsxPathInput.value = data.path;
        if (appState.project) { setProjectStatus('Unsaved changes'); scheduleProjectSave(); }
      }
    }

    async function updateModelOutputs() {
      if (!appState.project) { showError('Start or open a project first.'); return; }
      if (!memoPathInput.value.trim()) { showError('Point to a memo file first (Browse next to "Memo file").'); return; }
      if (!modelXlsxPathInput.value.trim()) { showError('Enter the Excel model filepath first.'); return; }
      const payload = {...projectPayloadFromState(), fund_infos: getFundInfos()};
      updateModelOutputsBtn.disabled = true;
      setProjectStatus('Updating model outputs in memo…');
      try {
        const resp = await fetch('/api/projects/update_model_outputs', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify(payload),
        });
        const data = await resp.json();
        if (!resp.ok) throw new Error(data.error || 'Unable to update model outputs.');
        setProjectStatus('Model outputs updated in: ' + (data.memo_file_path || 'memo'));
      } finally {
        updateModelOutputsBtn.disabled = false;
      }
    }

    async function refreshSheetPreview(uploadId, sheetName, headerMode) {
      const resp = await fetch('/api/preview', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ upload_id: uploadId, sheet_name: sheetName, header_mode: headerMode || 'auto' }),
      });
      const data = await resp.json();
      if (!resp.ok) {
        throw new Error(data.error || 'Unable to preview sheet.');
      }
      const target = appState.funds.find((item) => item.upload_id === uploadId);
      if (target) {
        target.sheet_name = sheetName;
        target.header_mode = data.header_mode || headerMode || 'auto';
        target.preview = data;
        target.mapping = data.suggested_columns;
      }
      renderFunds();
    }

    function renderFunds() {
      fundList.innerHTML = '';
      appState.funds.forEach((fund) => {
        const card = buildFundCard(fund);
        fundList.appendChild(card);
      });
    }

    function buildFundCard(fund) {
      const template = document.getElementById('fundCardTemplate');
      const node = template.content.firstElementChild.cloneNode(true);
      node.dataset.uploadId = fund.upload_id;

      const hasWorkbook = !!fund.preview;
      node.querySelector('.fund-name').textContent = fund.fund_name || fund.filename || 'Fund';
      node.querySelector('.fund-meta').textContent = hasWorkbook
        ? `${fund.filename || ''} | ${fund.row_count || 0} rows | ${(fund.sheet_names || []).length} sheet(s)`
        : 'No workbook imported (set a manual mix below, or import Excel)';
      const statusPill = node.querySelector('.fund-status');
      if (statusPill) statusPill.textContent = hasWorkbook ? 'workbook imported' : 'no workbook';

      const importInput = node.querySelector('.fund-import');
      if (importInput) {
        importInput.addEventListener('change', () => {
          const file = importInput.files && importInput.files[0];
          if (file) importFundExcel(fund, file).catch((err) => showError(err.message || String(err)));
        });
      }
      const importText = node.querySelector('.fund-import-text');
      if (importText) importText.textContent = hasWorkbook ? 'Replace Excel' : 'Import Excel';

      const removeBtn = node.querySelector('.fund-remove');
      if (removeBtn) removeBtn.addEventListener('click', () => removeFund(fund));

      const labelInput = node.querySelector('.fund-label');
      labelInput.value = fund.fund_name || defaultLabel(fund.filename || '');
      labelInput.addEventListener('input', () => {
        fund.fund_name = labelInput.value;
        scheduleProjectSave();
        scheduleChartRefresh();
      });

      const abbrevInput = node.querySelector('.fund-abbrev');
      if (abbrevInput) {
        abbrevInput.value = fund.abbrev_name || '';
        abbrevInput.addEventListener('input', () => {
          fund.abbrev_name = abbrevInput.value;
          scheduleProjectSave();
        });
      }

      const bidInput = node.querySelector('.fund-bid');
      bidInput.value = fund.bid_amount ?? '';
      bidInput.placeholder = '0';
      bidInput.addEventListener('input', () => {
        fund.bid_amount = bidInput.value ? Number(bidInput.value) : 0;
        scheduleProjectSave();
        scheduleChartRefresh();
      });

      const manualEditor = node.querySelector('.manual-editor');
      buildManualEditor(fund, manualEditor);

      const headerModeSelect = node.querySelector('.fund-header-mode');
      populateHeaderModeSelect(headerModeSelect, fund.header_mode || 'auto');
      headerModeSelect.addEventListener('change', async () => {
        const chosen = headerModeSelect.value;
        const target = appState.funds.find((item) => item.upload_id === fund.upload_id);
        if (!target) return;
        target.header_mode = chosen;
        try {
          await refreshSheetPreview(fund.upload_id, target.sheet_name || target.default_sheet, chosen);
          scheduleProjectSave();
          scheduleChartRefresh();
        } catch (err) {
          showError(err.message || String(err));
        }
      });

      const sheetSelect = node.querySelector('.fund-sheet');
      fund.sheet_names.forEach((sheet) => {
        const option = document.createElement('option');
        option.value = sheet;
        option.textContent = sheet;
        if (sheet === fund.default_sheet) option.selected = true;
        sheetSelect.appendChild(option);
      });
      sheetSelect.addEventListener('change', async () => {
        const chosen = sheetSelect.value;
        const target = appState.funds.find((item) => item.upload_id === fund.upload_id);
        if (!target) return;
        target.sheet_name = chosen;
        try {
          await refreshSheetPreview(fund.upload_id, chosen, target.header_mode || 'auto');
          scheduleProjectSave();
          scheduleChartRefresh();
        } catch (err) {
          showError(err.message || String(err));
        }
      });

      const mappingGrid = node.querySelector('.fund-mapping');
      const mappingFields = ['investment_name', 'record_date_nav', 'asset_class', 'security_type', 'geography'];
      let securityTypeSelect = null;
      mappingFields.forEach((field) => {
        const wrapper = document.createElement('div');
        wrapper.className = 'field';
        const label = document.createElement('label');
        label.textContent = FIELD_LABELS[field];
        const select = document.createElement('select');
        select.dataset.field = field;
        populateColumnSelect(select, fund.preview ? fund.preview.columns : [], (fund.mapping || {})[field]);
        select.addEventListener('change', () => {
          fund.mapping = fund.mapping || {};
          fund.mapping[field] = select.value || null;
          scheduleProjectSave();
          scheduleChartRefresh();
        });
        if (field === 'security_type') {
          securityTypeSelect = select;
        }
        wrapper.appendChild(label);
        wrapper.appendChild(select);
        mappingGrid.appendChild(wrapper);
      });

      const previewTable = node.querySelector('.preview-table');
      if (fund.preview) {
        renderPreviewTable(previewTable, fund.preview);
      }

      node._fund = fund;
      return node;
    }

    function populateColumnSelect(select, columns, selectedValue) {
      select.innerHTML = '';
      const empty = document.createElement('option');
      empty.value = '';
      empty.textContent = '(not mapped)';
      select.appendChild(empty);
      columns.forEach((column) => {
        const option = document.createElement('option');
        option.value = column;
        option.textContent = column;
        if (column === selectedValue) option.selected = true;
        select.appendChild(option);
      });
      if (selectedValue && !columns.includes(selectedValue)) {
        const option = document.createElement('option');
        option.value = selectedValue;
        option.textContent = selectedValue + ' (saved)';
        option.selected = true;
        select.appendChild(option);
      }
    }

    function renderPreviewTable(table, preview) {
      const rows = preview.sample_rows || [];
      const columns = preview.columns || [];
      if (!rows.length) {
        table.innerHTML = '<tr><td class="muted">No rows to preview.</td></tr>';
        return;
      }
      const head = '<thead><tr>' + columns.map((col) => `<th>${escapeHtml(col)}</th>`).join('') + '</tr></thead>';
      const body = '<tbody>' + rows.map((row) => {
        return '<tr>' + columns.map((col) => `<td>${escapeHtml(row[col] ?? '')}</td>`).join('') + '</tr>';
      }).join('') + '</tbody>';
      table.innerHTML = head + body;
    }

    function collectFundState(card) {
      const uploadId = card.dataset.uploadId;
      const fund = appState.funds.find((item) => item.upload_id === uploadId);
      const label = card.querySelector('.fund-label').value.trim();
      const abbrevName = (card.querySelector('.fund-abbrev')?.value || '').trim();
      const bidAmount = parseFloat(card.querySelector('.fund-bid').value || '0') || 0;
      const sheetName = card.querySelector('.fund-sheet').value;
      const columnMap = {};
      card.querySelectorAll('.fund-mapping select').forEach((select) => {
        columnMap[select.dataset.field] = select.value || null;
      });
      return {
        upload_id: uploadId,
        filename: fund.filename,
        fund_name: label || defaultLabel(fund.filename || '') || 'Fund',
        abbrev_name: abbrevName,
        sheet_name: sheetName,
        header_mode: card.querySelector('.fund-header-mode').value || 'auto',
        bid_amount: bidAmount,
        manual_category_overrides: fund.manual_category_overrides || {},
        column_map: columnMap,
      };
    }

    async function calculate() {
      showError('');
      if (!appState.funds.length) {
        showError('Load workbooks first.');
        return;
      }
      const funds = Array.from(document.querySelectorAll('.fund-card')).map(collectFundState);
      appState.calculating = true;
      calcBtn.disabled = true;
      calcBtn.textContent = 'Calculating...';
      try {
        const resp = await fetch('/api/calculate', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({
            funds,
            normalization_rules: rulesBox.value,
            issuer_aliases: aliasesBox.value,
          }),
        });
        const data = await resp.json();
        if (!resp.ok) {
          throw new Error(data.error || 'Unable to calculate exposures.');
        }
        appState.result = data;
        renderResults(data);
        autoFillLabelColors(data);
        renderFundInfoTable();
      } catch (err) {
        showError(err.message || String(err));
      } finally {
        appState.calculating = false;
        calcBtn.disabled = false;
        calcBtn.textContent = 'Calculate exposures';
      }
    }

    function renderResults(data) {
      const stats = [
        { k: 'Total bid', v: money(data.total_bid), s: `${data.funds.length} funds` },
        { k: 'Top 1 position', v: pct(data.top_concentration.top_1), s: 'project share' },
        { k: 'Top 3 positions', v: pct(data.top_concentration.top_3), s: 'project share' },
        { k: 'Top 5 positions', v: pct(data.top_concentration.top_5), s: 'project share' },
        { k: 'Top 10 positions', v: pct(data.top_concentration.top_10), s: 'project share' },
      ];
      const sentence = data.summary_sentence ? `<p class="small-note">${escapeHtml(data.summary_sentence)}</p>` : '';
      const statHtml = '<div class="stats">' + stats.map((stat) => `
        <div class="stat">
          <div class="k">${escapeHtml(stat.k)}</div>
          <div class="v">${escapeHtml(stat.v)}</div>
          <div class="s">${escapeHtml(stat.s)}</div>
        </div>
      `).join('') + '</div>';

      const familyOrder = [
        ['asset_class', 'Asset class exposure'],
        ['security_type', 'Security type exposure'],
        ['geography', 'Geography exposure'],
      ];
      const chartHtml = '<div class="charts">' + familyOrder
        .filter(([family]) => (data.categories[family] || []).length)
        .map(([family, title]) => renderCategoryCard(family, title, data.categories[family]))
        .join('') + '</div>';
      const perFundHtml = renderPerFundExposure(data.fund_profiles || [], data.categories || {}, data.top_positions || []);
      const topPositionsHtml = renderBarCard('Top positions', data.top_positions.slice(0, 15));
      const concentrationTableHtml = renderConcentrationTable(data);
      const mergesHtml = renderMergesCard(data.position_merges || [], data.position_merge_suggestions || []);
      const fundTableHtml = renderFundTable(data.funds);

      resultsPane.innerHTML = sentence + statHtml + '<div class="section-gap"></div>' + chartHtml + (perFundHtml ? '<div class="section-gap"></div>' + perFundHtml : '') + '<div class="section-gap"></div>' + concentrationTableHtml + '<div class="section-gap"></div>' + topPositionsHtml + '<div class="section-gap"></div>' + mergesHtml + '<div class="section-gap"></div>' + fundTableHtml;
    }

    function autoFillLabelColors(data) {
      // Only auto-fill if the box is empty — don't overwrite manual edits.
      if (labelColorsBox.value.trim()) return;
      const cats = data.categories || {};
      const lines = [];
      const THRESHOLD = 0.07;
      ['asset_class', 'security_type', 'geography'].forEach((fam) => {
        (cats[fam] || []).forEach((item) => {
          const pct = Number(item.value || item.percentage || 0);
          const color = pct >= THRESHOLD ? 'white' : 'black';
          lines.push(`${item.label} = ${color}`);
        });
      });
      // Deduplicate (same label may appear in multiple families)
      const seen = new Set();
      const unique = lines.filter((l) => { const k = l.split('=')[0].trim().toLowerCase(); if (seen.has(k)) return false; seen.add(k); return true; });
      labelColorsBox.value = unique.join('\n');
      scheduleProjectSave();
    }

    // ---- Fund Information ----

    const FI_FIELDS = ['lp_nav','unfunded','commits','invest_end','term_end','extensions','irr','tvpi','rvpi','dpi','leverage'];
    const FI_LABELS = ['LP NAV','Unf.','Commits','Invest End','Term End','Extensions','IRR','TVPI','RVPI','DPI','Leverage'];

    function getFundInfos() { return (appState.project && appState.project.fund_infos) ? appState.project.fund_infos : {}; }

    function saveFundInfos(fis) {
      if (!appState.project) return;
      appState.project.fund_infos = fis;
      setProjectStatus('Unsaved changes');
      scheduleProjectSave();
    }

    function getEffective(field) {
      if (!field || typeof field !== 'object') return field || '';
      return field.override !== null && field.override !== undefined ? field.override : (field.value ?? '');
    }

    function renderFundInfoTable() {
      const profiles = appState.result ? (appState.result.fund_profiles || []) : [];
      const fis = getFundInfos();
      if (!profiles.length) {
        fundInfoTableWrap.innerHTML = '<div class="small-note">Calculate exposures first to populate fund list.</div>';
        renderFundDocsList([], fis);
        return;
      }
      let html = `<div class="preview-wrap"><table><thead>
        <tr><th>Fund</th><th>Scale%</th>${FI_LABELS.map(l=>`<th>${escapeHtml(l)}</th>`).join('')}</tr>
      </thead><tbody>`;
      profiles.forEach((fp) => {
        const fname = fp.fund_name || fp.filename || 'Fund';
        const fi = fis[fname] || {};
        const fields = fi.fields || {};
        const scale = fi.scale_pct !== undefined ? fi.scale_pct : 100;
        html += `<tr data-fund="${escapeHtml(fname)}">
          <td>${escapeHtml(fname)}</td>
          <td><input type="number" class="fi-scale" data-fund="${escapeHtml(fname)}" value="${scale}" min="0" max="100" step="0.1" style="width:64px"/></td>`;
        FI_FIELDS.forEach((key) => {
          const f = fields[key] || {};
          const val = getEffective(f);
          const src = f.source || '';
          const perp = key === 'extensions' && f.perpetuity ? ' *' : '';
          const locked = !!f.locked;
          const lockCls = locked ? 'locked' : '';
          const lockTitle = locked ? 'Locked — click to unlock' : 'Lock to protect from parse overwrites';
          html += `<td>
            <div style="display:flex;align-items:center;gap:2px;">
              <input type="text" class="fi-cell ${lockCls}" data-fund="${escapeHtml(fname)}" data-field="${key}"
                value="${escapeHtml(String(val !== null && val !== undefined ? val : ''))}"
                placeholder="${src ? '(parsed)' : ''}" style="width:82px"/>
              <button class="fi-lock-btn ${lockCls}" data-fund="${escapeHtml(fname)}" data-field="${key}"
                title="${escapeHtml(lockTitle)}">${locked ? '🔒' : '🔓'}</button>
              ${src ? `<span title="${escapeHtml(src)}" style="cursor:help;color:var(--accent);font-size:11px">ⓘ</span>${perp}` : perp}
            </div>
          </td>`;
        });
        html += '</tr>';
      });
      html += '</tbody></table></div>';
      fundInfoTableWrap.innerHTML = html;
      // Listeners
      fundInfoTableWrap.querySelectorAll('.fi-cell').forEach((input) => {
        input.addEventListener('change', () => {
          const fname = input.dataset.fund;
          const field = input.dataset.field;
          const fis2 = getFundInfos();
          if (!fis2[fname]) fis2[fname] = {scale_pct:100, fields:{}};
          if (!fis2[fname].fields[field]) fis2[fname].fields[field] = {value:null,source:null,override:null};
          const v = input.value.trim();
          fis2[fname].fields[field].override = v === '' ? null : v;
          saveFundInfos(fis2);
        });
      });
      fundInfoTableWrap.querySelectorAll('.fi-lock-btn').forEach((btn) => {
        btn.addEventListener('click', () => {
          const fname = btn.dataset.fund;
          const field = btn.dataset.field;
          const fis2 = getFundInfos();
          if (!fis2[fname]) fis2[fname] = {scale_pct:100, fields:{}};
          if (!fis2[fname].fields[field]) fis2[fname].fields[field] = {value:null,source:null,override:null,locked:false};
          fis2[fname].fields[field].locked = !fis2[fname].fields[field].locked;
          saveFundInfos(fis2);
          renderFundInfoTable();
        });
      });
      fundInfoTableWrap.querySelectorAll('.fi-scale').forEach((input) => {
        input.addEventListener('change', () => {
          const fname = input.dataset.fund;
          const fis2 = getFundInfos();
          if (!fis2[fname]) fis2[fname] = {scale_pct:100, fields:{}};
          fis2[fname].scale_pct = parseFloat(input.value) || 100;
          saveFundInfos(fis2);
        });
      });
      renderFundDocsList(profiles, fis);
      wireDocDropZones();
    }

    function renderFundDocsList(profiles, fis) {
      if (!profiles.length) { fundDocsList.innerHTML = ''; return; }
      let html = '';
      profiles.forEach((fp) => {
        const fname = fp.fund_name || 'Fund';
        const fi = fis[fname] || {};
        const safeF = fname.replace(/"/g, '&quot;');
        const ql = fi.quarterly_letter_path || '';
        const afs = fi.afs_path || '';
        const cas = fi.cas_path || '';
        const lpa = fi.lpa_path || '';
        html += `<div class="chart-card" style="margin-bottom:10px;">
          <h3>${escapeHtml(fname)} — Documents</h3>
          <div class="grid" style="grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:8px;">
            ${docRow(fname,'quarterly_letter','Quarterly Letter', ql)}
            ${docRow(fname,'afs','Audited Financials (AFS)', afs)}
            ${docRow(fname,'cas','Capital Account Statement (CAS)', cas)}
            ${docRow(fname,'lpa','LPA', lpa)}
          </div>
        </div>`;
      });
      fundDocsList.innerHTML = html;
      wireDocDropZones();
      fundDocsList.querySelectorAll('.fi-doc-upload').forEach((inp) => {
        inp.addEventListener('change', () => uploadFundDoc(inp.dataset.fund, inp.dataset.dtype, inp.files[0]));
      });
      fundDocsList.querySelectorAll('.fi-parse-btn').forEach((btn) => {
        btn.addEventListener('click', () => parseFundDoc(btn.dataset.fund, btn.dataset.dtype, btn));
      });
    }

    function docRow(fname, dtype, label, currentPath) {
      const safeF = fname.replace(/"/g,'&quot;');
      const hasFile = !!currentPath;
      return `<div class="field">
        <label>${escapeHtml(label)}</label>
        <div class="drop-zone fi-drop-zone" data-fund="${safeF}" data-dtype="${dtype}" style="display:flex;flex-direction:column;gap:6px;">
          <div style="display:flex;gap:6px;align-items:center;flex-wrap:wrap;">
            <label class="btn secondary" style="font-size:12px;padding:6px 10px;white-space:nowrap;">Choose<input type="file" class="fi-doc-upload" data-fund="${safeF}" data-dtype="${dtype}" accept=".pdf" style="display:none"/></label>
            <input type="text" class="fi-doc-path" data-fund="${safeF}" data-dtype="${dtype}" value="${escapeHtml(currentPath)}" placeholder="Drag & drop or paste path…" style="flex:1;font-size:12px;min-width:120px"/>
            <button class="btn fi-parse-btn" data-fund="${safeF}" data-dtype="${dtype}" style="font-size:12px;padding:6px 10px;">${hasFile ? 'Re-parse' : 'Parse'}</button>
          </div>
          ${hasFile ? `<div class="small-note" style="font-size:11px;">${escapeHtml(currentPath.split(/[/\\]/).pop())}</div>` : ''}
        </div>
      </div>`;
    }

    async function uploadFundDoc(fname, dtype, file) {
      if (!file || !appState.project) return;
      showFundInfoError('');
      const form = new FormData();
      form.append('project_id', appState.project.project_id || '');
      form.append('fund_name', fname);
      form.append('doc_type', dtype);
      form.append('file', file);
      const r = await fetch('/api/fund_info/upload_doc', {method:'POST', body:form});
      const d = await r.json();
      if (!r.ok) { showFundInfoError(d.error || 'Upload failed'); return; }
      const fis = getFundInfos();
      if (!fis[fname]) fis[fname] = {scale_pct:100, fields:{}};
      fis[fname][dtype + '_path'] = d.path;
      saveFundInfos(fis);
      renderFundInfoTable();
      setProjectStatus(`Uploaded ${dtype} for ${fname}. Click Parse to extract data.`);
    }

    async function parseFundDoc(fname, dtype, btn) {
      if (!appState.project) return;
      const fis = getFundInfos();
      const fi = fis[fname] || {};
      // Check for path from path input first
      const pathInput = fundDocsList.querySelector(`.fi-doc-path[data-fund="${fname}"][data-dtype="${dtype}"]`);
      if (pathInput && pathInput.value.trim()) {
        if (!fi[dtype + '_path'] || fi[dtype + '_path'] !== pathInput.value.trim()) {
          if (!fis[fname]) fis[fname] = {scale_pct:100, fields:{}};
          fis[fname][dtype + '_path'] = pathInput.value.trim();
          saveFundInfos(fis);
        }
      }
      const path = (fis[fname] || {})[dtype + '_path'];
      if (!path) { showFundInfoError(`No ${dtype} file set for ${fname}. Upload or paste a path first.`); return; }
      const apiKey = openaiKeyInput.value.trim();
      if (!apiKey) { showFundInfoError('Enter your OpenAI API key first.'); return; }
      const origText = btn.textContent;
      btn.disabled = true; btn.textContent = 'Parsing…';
      const docLabel = {'quarterly_letter':'quarterly letter','afs':'AFS','cas':'CAS','lpa':'LPA'}[dtype] || dtype;
      showParseSpinner(`Parsing ${docLabel} for ${fname}…`);
      try {
        const r = await fetch('/api/fund_info/parse', {method:'POST',
          headers:{'Content-Type':'application/json'},
          body:JSON.stringify({project_id:appState.project.project_id,fund_name:fname,doc_type:dtype,file_path:path,api_key:apiKey})});
        const d = await r.json();
        if (!r.ok) { hideParseSpinner(); showFundInfoError(d.error || 'Parse failed'); return; }
        const fis2 = getFundInfos();
        if (!fis2[fname]) fis2[fname] = {scale_pct:100, fields:{}};
        mergeFields(fis2[fname], d.fields || {});
        saveFundInfos(fis2);
        renderFundInfoTable();
        const populated = Object.entries(d.fields || {}).filter(([,v])=>v&&v.value!=null).map(([k])=>k).join(', ');
        hideParseSpinner(`✓ Parsed ${docLabel} for ${fname}${populated ? ' — populated: ' + populated : ''}.`);
      } catch(e) { hideParseSpinner(); showFundInfoError(e.message || String(e)); }
      finally { btn.disabled = false; btn.textContent = origText; }
    }

    async function parseProjFiles() {
      if (!appState.project || !appState.result) return;
      const apiKey = openaiKeyInput.value.trim();
      if (!apiKey) { showFundInfoError('Enter your OpenAI API key first.'); return; }
      const paths = appState.project.project_file_paths || [];
      if (!paths.length) { showFundInfoError('Upload or add project files first.'); return; }
      const fundNames = (appState.result.fund_profiles || []).map(fp => fp.fund_name || fp.filename || '').filter(Boolean);
      parseProjFilesBtn.disabled = true; parseProjFilesBtn.textContent = 'Parsing…';
      showParseSpinner(`Scanning and parsing all fields for ${fundNames.length} fund(s)…`);
      try {
        const r = await fetch('/api/fund_info/parse_project', {method:'POST',
          headers:{'Content-Type':'application/json'},
          body:JSON.stringify({project_id:appState.project.project_id,file_paths:paths,fund_names:fundNames,api_key:apiKey})});
        const d = await r.json();
        if (!r.ok) { hideParseSpinner(); showFundInfoError(d.error || 'Parse failed'); return; }
        const fis = getFundInfos();
        let populated = 0;
        Object.entries(d.results || {}).forEach(([fname, extracted]) => {
          if (!fis[fname]) fis[fname] = {scale_pct:100, fields:{}};
          mergeFields(fis[fname], extracted || {});
          populated += Object.values(extracted || {}).filter(v => v && v.value != null).length;
        });
        saveFundInfos(fis);
        renderFundInfoTable();
        const filesNote = d.files_used && d.files_used.length ? ` from ${d.files_used.length} file(s): ${d.files_used.slice(0,3).join(', ')}${d.files_used.length>3?'…':''}` : '';
        hideParseSpinner(`✓ Project files parsed — ${populated} value(s) populated across ${fundNames.length} fund(s)${filesNote}.`);
      } catch(e) { hideParseSpinner(); showFundInfoError(e.message || String(e)); }
      finally { parseProjFilesBtn.disabled = false; parseProjFilesBtn.textContent = 'Parse project files'; }
    }

    async function updateFundInfo() {
      if (!appState.project) { showError('Start or open a project first.'); return; }
      if (!memoPathInput.value.trim()) { showError('Point to a memo file first.'); return; }
      const payload = {...projectPayloadFromState(), fund_infos: getFundInfos()};
      updateFundInfoBtn.disabled = true;
      setProjectStatus('Updating fund information in memo…');
      try {
        const r = await fetch('/api/projects/update_fund_info', {method:'POST',
          headers:{'Content-Type':'application/json'}, body:JSON.stringify(payload)});
        const d = await r.json();
        if (!r.ok) throw new Error(d.error || 'Update failed');
        setProjectStatus('Fund information updated in memo.');
      } finally { updateFundInfoBtn.disabled = false; }
    }

    async function updateAll() {
      if (!appState.project) { showError('Start or open a project first.'); return; }
      updateAllBtn.disabled = true;
      setProjectStatus('Updating all…');
      try {
        await updateExposures();
        await updateFundInfo();
        setProjectStatus('Update All complete.');
      } catch(e) { showError(e.message || String(e)); }
      finally { updateAllBtn.disabled = false; }
    }

    function renderConcentrationTable(data) {
      const profiles = data.fund_profiles || [];
      const top = data.top_concentration || {};
      function fundConc(fp) {
        const pos = fp.position_exposure || [];
        const w = Number(fp.weight || 0);
        const scale = w > 0 ? 1 / w : 1;   // convert project-share → fund-level %
        const vals = pos.map((p) => Number(p.value || p.percentage || 0) * scale).sort((a, b) => b - a);
        const sum = (n) => vals.slice(0, n).reduce((a, b) => a + b, 0);
        return { top_1: sum(1), top_3: sum(3), top_5: sum(5), top_10: sum(10), remaining: Math.max(0, 1 - sum(10)) };
      }
      const rows = [
        ['Total', top.top_1, top.top_3, top.top_5, top.top_10, top.remaining],
        ...profiles.map((fp) => {
          const c = fundConc(fp);
          return [fp.fund_name || fp.filename || 'Fund', c.top_1, c.top_3, c.top_5, c.top_10, c.remaining];
        }),
      ];
      const thead = '<thead><tr><th>Top Positions</th><th>Top 1</th><th>Top 3</th><th>Top 5</th><th>Top 10</th><th>Remaining</th></tr></thead>';
      const tbody = '<tbody>' + rows.map((r) =>
        '<tr>' + r.map((v, i) => `<td>${i === 0 ? escapeHtml(String(v)) : pct(v)}</td>`).join('') + '</tr>'
      ).join('') + '</tbody>';
      return `<div class="chart-card"><h3>Top Positions</h3><div class="preview-wrap"><table>${thead}${tbody}</table></div></div>`;
    }

    function unsplitMerge(variants) {
      const existing = (splitNamesBox.value || '').split('\n').map((s) => s.trim().toLowerCase()).filter(Boolean);
      let added = 0;
      (variants || []).forEach((v) => {
        if (!existing.includes(v.trim().toLowerCase())) {
          splitNamesBox.value = (splitNamesBox.value ? splitNamesBox.value + '\n' : '') + v.trim();
          added++;
        }
      });
      if (added > 0) {
        if (appState.project) { setProjectStatus('Unsaved changes'); scheduleProjectSave(); }
        scheduleChartRefresh();
      }
    }

    function renderMergesCard(merges, suggestions) {
      merges = merges || [];
      suggestions = suggestions || [];
      if (!merges.length && !suggestions.length) return '';
      let html = '<div class="chart-card"><h3>Combined names (review)</h3>';
      if (merges.length) {
        const rows = merges.map((m) => `
          <tr>
            <td>${escapeHtml(m.label)}</td>
            <td>${(m.variants || []).map((v) => escapeHtml(v)).join('<br>')}</td>
            <td><button class="btn" style="white-space:nowrap" onclick="unsplitMerge(${JSON.stringify(m.variants || [])})">Unsplit</button></td>
          </tr>
        `).join('');
        html += '<div class="small-note">These raw names were treated as the same issuer and combined at the project level. Click Unsplit to keep them separate.</div>';
        html += `<div class="preview-wrap"><table><thead><tr><th>Combined as</th><th>From these names</th><th></th></tr></thead><tbody>${rows}</tbody></table></div>`;
      }
      if (suggestions.length) {
        const items = suggestions.map((s) => `<li>${escapeHtml(s.a)} &nbsp;vs&nbsp; ${escapeHtml(s.b)}</li>`).join('');
        html += '<div class="small-note section-gap">Possibly the same issuer but <strong>not</strong> combined &mdash; if they should be, add an alias above (e.g. <code>' + (suggestions[0] ? escapeHtml(suggestions[0].a) + ' => ' + escapeHtml(suggestions[0].b) : 'variant => issuer') + '</code>):</div>';
        html += `<ul>${items}</ul>`;
      }
      html += '</div>';
      return html;
    }

    function categoryColor(family, label, index) {
      if (family === 'geography') {
        return GEOGRAPHY_COLORS[label] || ['#4f6fb5', '#c27c3d', '#8a4f69'][index % 3];
      }
      const assetClass = family === 'security_type'
        ? assetClassForSecurityType(label)
        : label;
      return ASSET_CLASS_COLORS[assetClass] || ['#1f5f74', '#8b5e3c', '#5d7d4e'][index % 3];
    }

    function renderCategoryCard(family, title, items) {
      const total = items.reduce((sum, item) => sum + Number(item.value || 0), 0) || 1;
      const size = 180;
      let angle = -90;
      const cx = size / 2;
      const cy = size / 2;
      const r = 70;
      const inner = 42;
      const slices = items.map((item, index) => {
        const value = Number(item.value || 0);
        const slice = value / total;
        const start = angle;
        const end = angle + slice * 360;
        angle = end;
        return donutSlice(cx, cy, r, inner, start, end, categoryColor(family, item.label, index));
      }).join('');
      const legend = items.map((item, index) => `
        <div class="legend-item">
          <span class="swatch" style="background:${categoryColor(family, item.label, index)}"></span>
          <span>${escapeHtml(item.label)}</span>
          <span>${pct(item.value)}</span>
        </div>
      `).join('');
      return `
        <div class="chart-card">
          <h3>${escapeHtml(title)}</h3>
          <div class="chart-row">
            <svg width="${size}" height="${size}" viewBox="0 0 ${size} ${size}" aria-label="${escapeHtml(title)}">
              ${slices}
              <circle cx="${cx}" cy="${cy}" r="${inner}" fill="#fff"></circle>
            </svg>
            <div class="legend">${legend}</div>
          </div>
        </div>
      `;
    }

    function renderPerFundExposure(fundProfiles, blendedCategories, topPositions) {
      if (!fundProfiles || fundProfiles.length === 0) return '';
      const totalProjectPositions = (topPositions || []).length;
      const posRows = fundProfiles.map(fp =>
        `<tr><td>${escapeHtml(fp.fund_name || fp.filename)}</td><td style="text-align:right">${pct(fp.weight)} of bid</td><td style="text-align:right">${fp.positions ?? '—'}</td></tr>`
      ).join('');
      const posCard = `<div class="chart-card">
        <h3>Positions by fund</h3>
        <div class="preview-wrap">
          <table>
            <thead><tr><th>Fund</th><th style="text-align:right">Bid weight</th><th style="text-align:right">Positions</th></tr></thead>
            <tbody>
              ${posRows}
              <tr style="font-weight:600;border-top:2px solid #e5e7eb">
                <td>Total (project)</td><td></td><td style="text-align:right">${totalProjectPositions}</td>
              </tr>
            </tbody>
          </table>
        </div>
      </div>`;
      const families = [
        ['asset_class', 'Asset class'],
        ['security_type', 'Security type'],
        ['geography', 'Geography'],
      ];
      const cards = families.map(([family, title]) => {
        // Use blended order so rows match the donut chart order above
        const blendedOrder = (blendedCategories[family] || []).map(i => i.label);
        const allLabels = new Set(blendedOrder);
        fundProfiles.forEach(fp => (fp.categories[family] || []).forEach(i => allLabels.add(i.label)));
        const extraLabels = Array.from(allLabels).filter(l => !blendedOrder.includes(l)).sort();
        const labels = [...blendedOrder, ...extraLabels];
        if (!labels.length) return '';
        const headers = '<th>Category</th>' + fundProfiles.map(fp =>
          `<th style="text-align:right">${escapeHtml(fp.fund_name || fp.filename)}<br><span style="font-weight:normal;color:#888;font-size:0.85em">${pct(fp.weight)} of bid</span></th>`
        ).join('');
        const rows = labels.map(label => {
          const cells = fundProfiles.map(fp => {
            const item = (fp.categories[family] || []).find(i => i.label === label);
            return `<td style="text-align:right">${item ? pct(item.percentage) : '<span style="color:#ccc">—</span>'}</td>`;
          }).join('');
          return `<tr><td>${escapeHtml(label)}</td>${cells}</tr>`;
        }).join('');
        return `<div class="chart-card">
          <h3>${escapeHtml(title)} by fund</h3>
          <div class="preview-wrap">
            <table>
              <thead><tr>${headers}</tr></thead>
              <tbody>${rows}</tbody>
            </table>
          </div>
        </div>`;
      }).filter(Boolean);
      return '<div class="charts">' + posCard + (cards.length ? cards.join('') : '') + '</div>';
    }

    function donutSlice(cx, cy, r, inner, startDeg, endDeg, color) {
      const start = polar(cx, cy, r, endDeg);
      const end = polar(cx, cy, r, startDeg);
      const innerStart = polar(cx, cy, inner, startDeg);
      const innerEnd = polar(cx, cy, inner, endDeg);
      const largeArc = endDeg - startDeg <= 180 ? 0 : 1;
      return `
        <path d="M ${start.x} ${start.y}
                 A ${r} ${r} 0 ${largeArc} 0 ${end.x} ${end.y}
                 L ${innerStart.x} ${innerStart.y}
                 A ${inner} ${inner} 0 ${largeArc} 1 ${innerEnd.x} ${innerEnd.y}
                 Z" fill="${color}"></path>`;
    }

    function polar(cx, cy, r, deg) {
      const rad = (deg - 90) * Math.PI / 180.0;
      return { x: cx + r * Math.cos(rad), y: cy + r * Math.sin(rad) };
    }

    function renderBarCard(title, items) {
      if (!items.length) {
        return '';
      }
      const max = Math.max(...items.map((item) => Number(item.value || 0)), 0) || 1;
      const bars = items.map((item) => `
        <div class="bar-item">
          <div>${escapeHtml(item.label)}</div>
          <div class="bar-track"><div class="bar-fill" style="width:${Math.max(0, Number(item.value || 0) / max * 100)}%"></div></div>
          <div style="text-align:right">${pct(item.value)}</div>
        </div>
      `).join('');
      return `
        <div class="chart-card">
          <h3>${escapeHtml(title)}</h3>
          <div class="bar-list">${bars}</div>
        </div>
      `;
    }

    function renderFundTable(funds) {
      const rows = funds.map((fund) => `
        <tr>
          <td>${escapeHtml(fund.fund_name || fund.filename)}</td>
          <td>${money(fund.bid_amount)}</td>
          <td>${pct(fund.weight)}</td>
          <td>${money(fund.total_nav)}</td>
          <td>${fund.positions}</td>
          <td>${escapeHtml(fund.header_mode || 'auto')}</td>
          <td>${fund.cash_rows ?? 0}</td>
        </tr>
      `).join('');
      return `
        <div class="chart-card">
          <h3>Fund summary</h3>
          <div class="preview-wrap">
            <table>
              <thead><tr><th>Fund</th><th>Bid</th><th>Weight</th><th>Invested NAV</th><th>Positions</th><th>Header mode</th><th>Cash rows</th></tr></thead>
              <tbody>${rows}</tbody>
            </table>
          </div>
        </div>
      `;
    }

    function money(value) {
      const num = Number(value || 0);
      if (!isFinite(num)) return '-';
      return num.toLocaleString(undefined, { maximumFractionDigits: 2, minimumFractionDigits: 0 });
    }

    function pct(value) {
      const num = Number(value || 0);
      if (!isFinite(num)) return '-';
      return (num * 100).toFixed(1) + '%';
    }

    function downloadJson() {
      if (!appState.result) {
        showError('Run a calculation first.');
        return;
      }
      const blob = new Blob([JSON.stringify(appState.result, null, 2)], { type: 'application/json' });
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url;
      a.download = 'project_balance_exposure.json';
      a.click();
      URL.revokeObjectURL(url);
    }

    async function downloadPositionsExcel() {
      if (!appState.result) { showError('Run a calculation first.'); return; }
      const payload = { result: appState.result, project_name: appState.project ? (appState.project.project_name || 'Project') : 'Project' };
      const resp = await fetch('/api/download_positions_excel', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(payload),
      });
      if (!resp.ok) { const d = await resp.json().catch(() => ({})); showError(d.error || 'Unable to generate Excel.'); return; }
      const blob = await resp.blob();
      const name = (resp.headers.get('Content-Disposition') || '').match(/filename="?([^";]+)"?/i);
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url; a.download = name ? name[1] : 'position_exposure.xlsx';
      a.click(); URL.revokeObjectURL(url);
    }

    function csvCell(value) {
      const text = String(value ?? '');
      if (/[",\n]/.test(text)) {
        return '"' + text.replaceAll('"', '""') + '"';
      }
      return text;
    }

    clearBtn.addEventListener('click', () => {
      appState.funds = [];
      appState.result = null;
      fundList.innerHTML = '';
      resultsPane.innerHTML = '<div class="small-note">No exposure data yet. Add a fund and click calculate.</div>';
      showError('');
      scheduleProjectSave();
    });

    projectSelect.addEventListener('change', async () => {
      try {
        await openProject(projectSelect.value);
      } catch (err) {
        showError(err.message || String(err));
      }
    });
    newProjectBtn.addEventListener('click', async () => {
      try {
        await startNewProject();
      } catch (err) {
        showError(err.message || String(err));
      }
    });
    saveProjectBtn.addEventListener('click', async () => {
      try {
        await saveProject();
      } catch (err) {
        showError(err.message || String(err));
      }
    });
    exportMemoBtn.addEventListener('click', async () => {
      exportMemo().catch((err) => showError(err.message || String(err)));
    });
    projectNameInput.addEventListener('input', () => {
      if (appState.project) {
        appState.project.project_name = projectNameInput.value.trim() || 'Untitled project';
        setProjectStatus('Unsaved changes');
        scheduleProjectSave();
      }
    });
    memoNameInput.addEventListener('input', () => {
      if (appState.project) {
        appState.project.memo_name = memoNameInput.value.trim() || '';
        setProjectStatus('Unsaved changes');
        scheduleProjectSave();
      }
    });
    rulesBox.addEventListener('input', () => {
      if (appState.project) {
        setProjectStatus('Unsaved changes');
        scheduleProjectSave();
      }
      scheduleChartRefresh();
    });
    aliasesBox.addEventListener('input', () => {
      if (appState.project) {
        setProjectStatus('Unsaved changes');
        scheduleProjectSave();
      }
      scheduleChartRefresh();
    });
    labelColorsBox.addEventListener('input', () => {
      if (appState.project) {
        setProjectStatus('Unsaved changes');
        scheduleProjectSave();
      }
    });
    splitNamesBox.addEventListener('input', () => {
      if (appState.project) {
        setProjectStatus('Unsaved changes');
        scheduleProjectSave();
      }
      scheduleChartRefresh();
    });

    liveUpdateToggle.addEventListener('change', () => {
      if (liveUpdateToggle.checked) {
        scheduleChartRefresh();
      }
    });

    addFundBtn.addEventListener('click', addFund);
    updateExposuresBtn.addEventListener('click', () => updateExposures().catch((err) => showError(err.message || String(err))));
    updateFundInfoBtn.addEventListener('click', () => updateFundInfo().catch((err) => showError(err.message || String(err))));
    updateModelOutputsBtn.addEventListener('click', () => updateModelOutputs().catch((err) => showError(err.message || String(err))));
    updateAllBtn.addEventListener('click', () => updateAll().catch((err) => showError(err.message || String(err))));
    parseProjFilesBtn.addEventListener('click', () => parseProjFiles().catch((err) => showError(err.message || String(err))));
    openaiKeyInput.addEventListener('input', () => { if (appState.project) { setProjectStatus('Unsaved changes'); scheduleProjectSave(); } });
    projFilesPathInput.addEventListener('input', () => { if (appState.project) { setProjectStatus('Unsaved changes'); scheduleProjectSave(); } });
    projFileUpload.addEventListener('change', async () => {
      if (!appState.project || !projFileUpload.files.length) return;
      const form = new FormData();
      form.append('project_id', appState.project.project_id || '');
      Array.from(projFileUpload.files).forEach(f => form.append('files', f));
      const r = await fetch('/api/fund_info/upload_proj_files', {method:'POST', body:form});
      const d = await r.json();
      if (!r.ok) { showError(d.error || 'Upload failed'); return; }
      if (!appState.project.project_file_paths) appState.project.project_file_paths = [];
      (d.paths || []).forEach(p => { if (!appState.project.project_file_paths.includes(p)) appState.project.project_file_paths.push(p); });
      projFilesPathInput.value = appState.project.project_file_paths.join('\n');
      projFilesList.textContent = `${appState.project.project_file_paths.length} file(s) set.`;
      scheduleProjectSave();
    });
    browseMemoBtn.addEventListener('click', () => browseMemo().catch((err) => showError(err.message || String(err))));
    memoPathInput.addEventListener('input', () => {
      if (appState.project) { setProjectStatus('Unsaved changes'); scheduleProjectSave(); }
    });
    browseModelBtn.addEventListener('click', () => browseModel().catch((err) => showError(err.message || String(err))));
    modelXlsxPathInput.addEventListener('input', () => {
      if (appState.project) { setProjectStatus('Unsaved changes'); scheduleProjectSave(); }
    });
    calcBtn.addEventListener('click', calculate);
    downloadJsonBtn.addEventListener('click', downloadJson);
    downloadExcelBtn.addEventListener('click', () => downloadPositionsExcel().catch((err) => showError(err.message || String(err))));

    (async () => {
      try {
        await loadProjectList();
        if (appState.projects.length) {
          await openProject(appState.projects[0].project_id);
        } else {
          await startNewProject();
        }
        scheduleChartRefresh();
      } catch (err) {
        showError(err.message || String(err));
      }
    })();

    window.projectBalanceApp = {
      state: appState,
      renderFunds,
      renderResults,
      showError,
    };
    document.body.dataset.scriptReady = '1';
  </script>
</body>
</html>
"""


class Handler(BaseHTTPRequestHandler):
    def _send(self, status: int, content_type: str, body: bytes):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _send_json(self, status: int, payload: Dict[str, Any]):
        self._send(status, "application/json; charset=utf-8", json.dumps(payload).encode("utf-8"))

    def do_GET(self):
        if self.path == "/":
            self._send(200, "text/html; charset=utf-8", HTML.encode("utf-8"))
            return
        if self.path == "/api/projects":
            self._send_json(200, {"projects": list_project_summaries()})
            return
        self._send_json(404, {"error": "Not found"})

    def do_POST(self):
        try:
            if self.path == "/api/load":
                self._handle_load()
                return
            if self.path == "/api/load_base64":
                self._handle_load_base64()
                return
            if self.path == "/api/projects/new":
                self._handle_project_new()
                return
            if self.path == "/api/projects/open":
                self._handle_project_open()
                return
            if self.path == "/api/projects/save":
                self._handle_project_save()
                return
            if self.path == "/api/projects/export_memo":
                self._handle_project_export_memo()
                return
            if self.path == "/api/projects/update_exposures":
                self._handle_project_update_exposures()
                return
            if self.path == "/api/pick_file":
                self._handle_pick_file()
                return
            if self.path == "/api/preview":
                self._handle_preview()
                return
            if self.path == "/api/calculate":
                self._handle_calculate()
                return
            if self.path == "/api/download_positions_excel":
                self._handle_download_positions_excel()
                return
            if self.path == "/api/fund_info/upload_doc":
                self._handle_fund_info_upload_doc()
                return
            if self.path == "/api/fund_info/upload_proj_files":
                self._handle_fund_info_upload_proj_files()
                return
            if self.path == "/api/fund_info/parse":
                self._handle_fund_info_parse()
                return
            if self.path == "/api/fund_info/parse_project":
                self._handle_fund_info_parse_project()
                return
            if self.path == "/api/projects/update_fund_info":
                self._handle_project_update_fund_info()
                return
            if self.path == "/api/projects/update_model_outputs":
                self._handle_model_outputs_update()
                return
            self._send_json(404, {"error": "Not found"})
        except Exception as exc:
            self._send_json(400, {"error": str(exc)})

    def _handle_load(self):
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            },
        )
        items = form["files"]
        if not isinstance(items, list):
            items = [items]
        project_id = form.getfirst("project_id") or ""
        loaded = self._ingest_files(
            [
                {"name": item.filename or "workbook.xlsx", "data": item.file.read()}
                for item in items
            ],
            project_id=project_id,
        )
        self._send_json(200, {"funds": loaded})

    def _handle_load_base64(self):
        payload = self._read_json()
        files = payload.get("files") or []
        project_id = payload.get("project_id") or ""
        loaded = self._ingest_files(
            [
                {
                    "name": f.get("name") or "workbook.xlsx",
                    "data": base64.b64decode(f.get("base64") or ""),
                }
                for f in files
            ],
            project_id=project_id,
        )
        self._send_json(200, {"funds": loaded})

    def _handle_project_new(self):
        content_type = self.headers.get("Content-Type", "")
        project_name = ""
        memo_name = ""
        memo_file = None
        if content_type.startswith("multipart/form-data"):
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": content_type,
                },
            )
            project_name = form.getfirst("project_name") or ""
            memo_name = form.getfirst("memo_name") or ""
            memo_field = form["memo_file"] if "memo_file" in form else None
            if memo_field is not None and getattr(memo_field, "filename", ""):
                memo_file = {
                    "filename": memo_field.filename,
                    "data": memo_field.file.read(),
                }
        else:
            payload = self._read_json()
            project_name = payload.get("project_name") or ""
            memo_name = payload.get("memo_name") or ""

        project = create_project(project_name=project_name, memo_name=memo_name)
        if memo_file:
            memo_folder = project_dir(project["project_id"]) / "memo"
            memo_folder.mkdir(parents=True, exist_ok=True)
            memo_name_safe = Path(memo_file["filename"]).name
            memo_path = memo_folder / memo_name_safe
            memo_path.write_bytes(memo_file["data"])
            project["memo_file_name"] = memo_name_safe
            project["memo_file_path"] = str(memo_path)
            project = save_project(project)
        self._send_json(200, {"project": project, "projects": list_project_summaries()})

    def _handle_project_open(self):
        payload = self._read_json()
        project_id = payload.get("project_id")
        if not project_id:
            raise ValueError("Project id is required")
        project = load_project(project_id)
        funds = self._rehydrate_project_funds(project)
        self._send_json(200, {"project": project, "funds": funds, "projects": list_project_summaries()})

    def _handle_project_save(self):
        payload = self._read_json()
        project_id = payload.get("project_id")
        if not project_id:
            raise ValueError("Project id is required")
        project = load_project(project_id)
        project["project_name"] = payload.get("project_name") or project.get("project_name") or "Untitled project"
        project["memo_name"] = payload.get("memo_name") or ""
        project["rules"] = payload.get("rules") or ""
        if payload.get("issuer_aliases") is not None:
            project["issuer_aliases"] = payload.get("issuer_aliases") or ""
        if payload.get("label_colors") is not None:
            project["label_colors"] = payload.get("label_colors") or ""
        if payload.get("force_splits") is not None:
            project["force_splits"] = payload.get("force_splits") or ""
        if payload.get("openai_key") is not None:
            project["openai_key"] = payload.get("openai_key") or ""
        if payload.get("project_file_paths") is not None:
            project["project_file_paths"] = payload.get("project_file_paths") or []
        if payload.get("fund_infos") is not None:
            project["fund_infos"] = payload.get("fund_infos") or {}
        if payload.get("memo_file_path") is not None:
            project["memo_file_path"] = payload.get("memo_file_path") or ""
        if payload.get("model_xlsx_path") is not None:
            project["model_xlsx_path"] = payload.get("model_xlsx_path") or ""
        payload_funds = self._normalize_funds_payload(payload.get("funds"))
        project["funds"] = self._persist_project_funds(project_id, payload_funds, project)
        saved = save_project(project)
        self._send_json(200, {"project": saved, "projects": list_project_summaries()})

    def _handle_project_update_exposures(self):
        payload = self._read_json()
        project_id = payload.get("project_id")
        if not project_id:
            raise ValueError("Project id is required")
        project = load_project(project_id)
        project["project_name"] = payload.get("project_name") or project.get("project_name") or "Untitled project"
        project["memo_name"] = payload.get("memo_name") or project.get("memo_name") or ""
        project["rules"] = payload.get("rules") or project.get("rules") or ""
        if payload.get("issuer_aliases") is not None:
            project["issuer_aliases"] = payload.get("issuer_aliases") or ""
        if payload.get("label_colors") is not None:
            project["label_colors"] = payload.get("label_colors") or ""
        if payload.get("force_splits") is not None:
            project["force_splits"] = payload.get("force_splits") or ""
        if payload.get("openai_key") is not None:
            project["openai_key"] = payload.get("openai_key") or ""
        if payload.get("project_file_paths") is not None:
            project["project_file_paths"] = payload.get("project_file_paths") or []
        if payload.get("fund_infos") is not None:
            project["fund_infos"] = payload.get("fund_infos") or {}
        if payload.get("memo_file_path") is not None:
            project["memo_file_path"] = payload.get("memo_file_path") or ""
        memo_path = (project.get("memo_file_path") or "").strip().strip('"').strip("'")
        if not memo_path:
            raise ValueError("Point to a memo file first (use Browse next to 'Memo file').")
        if not Path(memo_path).exists():
            raise ValueError(f"Memo file not found: {memo_path}")
        payload_funds = self._normalize_funds_payload(payload.get("funds"))
        project["funds"] = self._persist_project_funds(project_id, payload_funds, project)
        saved = save_project(project)
        result = compute_project_exposure(payload_funds, UPLOADS, parse_mapping_rules(project.get("rules") or ""), issuer_aliases=project.get("issuer_aliases") or "", force_splits=project.get("force_splits") or "")
        try:
            update_sections_in_file(memo_path, result, sections=("exposures", "portfolio_names"), project=saved)
        except PermissionError:
            raise ValueError("Could not write the memo - is it open in Word? Close it and try again.")
        self._send_json(200, {"ok": True, "memo_file_path": memo_path, "project": saved})

    def _handle_pick_file(self):
        # The server runs on the user's own machine, so we can pop a native Open dialog
        # to capture a real absolute path (a browser file input can't expose one).
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            path = filedialog.askopenfilename(
                title="Select the memo file (in your SharePoint/OneDrive synced folder)",
                filetypes=[("Word documents", "*.docx"), ("All files", "*.*")],
            )
            root.destroy()
        except Exception as exc:  # tkinter may be unavailable; UI falls back to pasting a path
            self._send_json(200, {"path": "", "error": f"File dialog unavailable ({exc}). Paste the path instead."})
            return
        self._send_json(200, {"path": path or ""})

    def _handle_project_export_memo(self):
        payload = self._read_json()
        project_id = payload.get("project_id")
        if not project_id:
            raise ValueError("Project id is required")
        project = load_project(project_id)
        project["project_name"] = payload.get("project_name") or project.get("project_name") or "Untitled project"
        project["memo_name"] = payload.get("memo_name") or project.get("memo_name") or ""
        project["rules"] = payload.get("rules") or project.get("rules") or ""
        if payload.get("issuer_aliases") is not None:
            project["issuer_aliases"] = payload.get("issuer_aliases") or ""
        if payload.get("label_colors") is not None:
            project["label_colors"] = payload.get("label_colors") or ""
        if payload.get("force_splits") is not None:
            project["force_splits"] = payload.get("force_splits") or ""
        if payload.get("openai_key") is not None:
            project["openai_key"] = payload.get("openai_key") or ""
        if payload.get("project_file_paths") is not None:
            project["project_file_paths"] = payload.get("project_file_paths") or []
        if payload.get("fund_infos") is not None:
            project["fund_infos"] = payload.get("fund_infos") or {}
        payload_funds = self._normalize_funds_payload(payload.get("funds"))
        project["funds"] = self._persist_project_funds(project_id, payload_funds, project)
        saved = save_project(project)
        result = compute_project_exposure(payload_funds, UPLOADS, parse_mapping_rules(project.get("rules") or ""), issuer_aliases=project.get("issuer_aliases") or "", force_splits=project.get("force_splits") or "")
        export_dir = project_dir(project_id) / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{(saved.get('project_name') or 'memo_helper').strip().replace(' ', '_')}_updated.docx"
        output_path = export_dir / filename
        build_memo_export(saved, result, output_path)
        content = output_path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        self.send_header("Content-Length", str(len(content)))
        self.send_header("Content-Disposition", f'attachment; filename="{output_path.name}"')
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(content)

    def _normalize_funds_payload(self, funds: Any) -> List[Dict[str, Any]]:
        if not funds:
            return []
        if isinstance(funds, str):
            try:
                funds = json.loads(funds)
            except Exception:
                return []
        if isinstance(funds, dict):
            funds = [funds]
        if not isinstance(funds, list):
            return []
        return [item for item in funds if isinstance(item, dict)]

    def _ingest_files(self, files: List[Dict[str, Any]], project_id: str = ""):
        loaded = []
        for item in files:
            filename = Path(item["name"]).name
            data = item["data"]
            upload_id = str(uuid.uuid4())
            sheet_names = read_workbook_metadata(data)
            if not sheet_names:
                raise ValueError(f"{filename} has no readable worksheets")
            default_sheet = sheet_names[0]
            preview = infer_sheet_preview(data, default_sheet)
            source_path = ""
            if project_id:
                source_path = store_upload_bytes(project_id, upload_id, filename, data)
            UPLOADS[upload_id] = {
                "filename": filename,
                "data": data,
                "sheet_names": sheet_names,
                "default_sheet": default_sheet,
                "header_mode": preview["header_mode"],
                "project_id": project_id or "",
                "source_path": source_path,
            }
            loaded.append(
                {
                    "upload_id": upload_id,
                    "filename": filename,
                    "sheet_names": sheet_names,
                    "default_sheet": default_sheet,
                    "header_mode": preview["header_mode"],
                    "row_count": preview["row_count"],
                    "preview": {
                        "columns": preview["columns"],
                        "sample_rows": preview["sample_rows"],
                    },
                    "mapping": {
                        **preview["suggested_columns"],
                    },
                }
            )
        return loaded

    def _persist_project_funds(self, project_id: str, funds: List[Dict[str, Any]], project: Dict[str, Any]):
        saved_funds: List[Dict[str, Any]] = []
        project_lookup = {item.get("upload_id"): item for item in project.get("funds") or []}
        for fund in funds:
            upload_id = fund.get("upload_id") or str(uuid.uuid4())
            upload = UPLOADS.get(upload_id)
            if upload and upload.get("project_id") == project_id and upload.get("source_path"):
                source_path = upload["source_path"]
            else:
                source_path = (project_lookup.get(upload_id) or {}).get("source_path") or ""
                if upload and project_id and not source_path:
                    source_path = store_upload_bytes(project_id, upload_id, upload["filename"], upload["data"])
                    upload["source_path"] = source_path
                    upload["project_id"] = project_id
            saved_funds.append(
                {
                    "upload_id": upload_id,
                    "filename": fund.get("filename") or (upload or {}).get("filename") or "workbook.xlsx",
                    "fund_name": fund.get("fund_name") or fund.get("filename") or "Untitled fund",
                    "sheet_name": fund.get("sheet_name") or (upload or {}).get("default_sheet") or "",
                    "header_mode": fund.get("header_mode") or (upload or {}).get("header_mode") or "auto",
                    "bid_amount": float(fund.get("bid_amount") or 0.0),
                    "abbrev_name": fund.get("abbrev_name") or "",
                    "manual_category_overrides": fund.get("manual_category_overrides") or {},
                    "column_map": fund.get("column_map") or {},
                    "source_path": source_path,
                }
            )
        return saved_funds

    def _rehydrate_project_funds(self, project: Dict[str, Any]):
        reloaded: List[Dict[str, Any]] = []
        project_id = project.get("project_id")
        for fund in project.get("funds") or []:
            upload_id = fund.get("upload_id") or str(uuid.uuid4())
            source_path = fund.get("source_path") or ""
            if not source_path:
                # Manual / not-yet-imported fund: no workbook to reload, keep its inputs.
                reloaded.append({
                    "upload_id": upload_id,
                    "sheet_names": [],
                    "default_sheet": "",
                    "sheet_name": fund.get("sheet_name") or "",
                    "header_mode": fund.get("header_mode") or "auto",
                    "row_count": 0,
                    "preview": None,
                    "mapping": fund.get("column_map") or {},
                    "fund_name": fund.get("fund_name") or "Untitled fund",
                    "bid_amount": fund.get("bid_amount") or 0,
                    "abbrev_name": fund.get("abbrev_name") or "",
                    "manual_category_overrides": fund.get("manual_category_overrides") or {},
                })
                continue
            data = read_upload_bytes(source_path, project_id=project_id)
            sheet_names = read_workbook_metadata(data)
            if not sheet_names:
                raise ValueError(f"{fund.get('filename') or 'workbook'} has no readable worksheets")
            sheet_name = fund.get("sheet_name") or sheet_names[0]
            header_mode = fund.get("header_mode") or "auto"
            preview = infer_sheet_preview(data, sheet_name, header_mode=header_mode)
            UPLOADS[upload_id] = {
                "filename": fund.get("filename") or Path(source_path).name,
                "data": data,
                "sheet_names": sheet_names,
                "default_sheet": sheet_names[0],
                "header_mode": preview["header_mode"],
                "project_id": project_id or "",
                "source_path": source_path,
            }
            reloaded.append(
                {
                    "upload_id": upload_id,
                    "filename": fund.get("filename") or Path(source_path).name,
                    "sheet_names": sheet_names,
                    "default_sheet": sheet_names[0],
                    "sheet_name": sheet_name,
                    "header_mode": preview["header_mode"],
                    "row_count": preview["row_count"],
                    "preview": {
                        "columns": preview["columns"],
                        "sample_rows": preview["sample_rows"],
                    },
                    "mapping": fund.get("column_map") or {
                        **preview["suggested_columns"],
                    },
                    "fund_name": fund.get("fund_name") or fund.get("filename") or "Untitled fund",
                    "bid_amount": fund.get("bid_amount") or 0,
                    "abbrev_name": fund.get("abbrev_name") or "",
                    "manual_category_overrides": fund.get("manual_category_overrides") or {},
                }
            )
        return reloaded

    def _handle_preview(self):
        payload = self._read_json()
        upload_id = payload.get("upload_id")
        sheet_name = payload.get("sheet_name")
        upload = UPLOADS.get(upload_id)
        if not upload:
            raise ValueError("Upload not found")
        if not sheet_name:
            sheet_name = upload["default_sheet"]
        header_mode = payload.get("header_mode") or upload.get("header_mode") or "auto"
        preview = infer_sheet_preview(upload["data"], sheet_name, header_mode=header_mode)
        upload["header_mode"] = preview["header_mode"]
        self._send_json(
            200,
            {
                "columns": preview["columns"],
                "sample_rows": preview["sample_rows"],
                "suggested_columns": preview["suggested_columns"],
                "row_count": preview["row_count"],
                "header_mode": preview["header_mode"],
            },
        )

    def _handle_fund_info_upload_doc(self):
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers,
                                environ={"REQUEST_METHOD":"POST","CONTENT_TYPE":self.headers.get("Content-Type")})
        project_id = form.getfirst("project_id") or ""
        fund_name  = form.getfirst("fund_name") or ""
        doc_type   = form.getfirst("doc_type") or ""
        f = form["file"] if "file" in form else None
        if not f or not getattr(f, "filename", None):
            raise ValueError("No file provided")
        from project_store import project_dir as _pdir
        dest_dir = _pdir(project_id) / "fund_docs" if project_id else Path("/tmp")
        dest_dir.mkdir(parents=True, exist_ok=True)
        safe = Path(f.filename).name
        dest = dest_dir / f"{fund_name}_{doc_type}_{safe}".replace(" ", "_")
        dest.write_bytes(f.file.read())
        self._send_json(200, {"path": str(dest)})

    def _handle_fund_info_upload_proj_files(self):
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers,
                                environ={"REQUEST_METHOD":"POST","CONTENT_TYPE":self.headers.get("Content-Type")})
        project_id = form.getfirst("project_id") or ""
        items = form["files"] if "files" in form else []
        if not isinstance(items, list): items = [items]
        from project_store import project_dir as _pdir
        dest_dir = _pdir(project_id) / "project_docs" if project_id else Path("/tmp")
        dest_dir.mkdir(parents=True, exist_ok=True)
        paths = []
        for item in items:
            if not getattr(item, "filename", None): continue
            dest = dest_dir / Path(item.filename).name
            dest.write_bytes(item.file.read())
            paths.append(str(dest))
        self._send_json(200, {"paths": paths})

    @staticmethod
    def _resolve_file_path(raw: str) -> Path:
        """Strip surrounding quotes and validate the path is readable."""
        p = Path(raw.strip().strip('"').strip("'"))
        if not p.exists():
            hint = (
                " It looks like a cloud-only OneDrive file — right-click it in "
                "Explorer → 'Always keep on this device', then retry."
                if "onedrive" in str(p).lower() else ""
            )
            raise ValueError(f"File not found: {p.name}.{hint}")
        try:
            with open(str(p), "rb") as fh:
                fh.read(4)
        except OSError:
            raise ValueError(
                f"Cannot open {p.name} — it may be a cloud-only OneDrive file. "
                "Right-click it in Explorer → 'Always keep on this device', then retry."
            )
        return p

    def _handle_fund_info_parse(self):
        import fund_info as fi_mod
        payload = self._read_json()
        fund_name = payload.get("fund_name") or ""
        doc_type  = payload.get("doc_type") or ""
        file_path = payload.get("file_path") or ""
        api_key   = payload.get("api_key") or ""
        if not file_path:
            raise ValueError("No file path provided")
        resolved = self._resolve_file_path(file_path)
        if not api_key:
            raise ValueError("OpenAI API key required")
        rpath = str(resolved)
        if doc_type == "quarterly_letter":
            result = fi_mod.parse_quarterly_letter(rpath, api_key, fund_name)
        elif doc_type == "afs":
            result = fi_mod.parse_afs(rpath, api_key, fund_name)
        elif doc_type == "cas":
            result = fi_mod.parse_cas(rpath, api_key, fund_name)
        elif doc_type == "lpa":
            result = fi_mod.parse_lpa(rpath, api_key, fund_name)
        else:
            raise ValueError(f"Unknown doc_type: {doc_type}")
        if "error" in result:
            raise ValueError(result["error"])
        self._send_json(200, {"fields": result})

    def _handle_fund_info_parse_project(self):
        import fund_info as fi_mod
        payload = self._read_json()
        file_paths = payload.get("file_paths") or []
        fund_names = payload.get("fund_names") or []
        api_key    = payload.get("api_key") or ""
        if not file_paths: raise ValueError("No project files provided")
        if not api_key:    raise ValueError("OpenAI API key required")
        raw = fi_mod.parse_project_files(file_paths, api_key, fund_names)
        files_used = raw.pop("_files_used", [])
        self._send_json(200, {"results": raw, "files_used": files_used})

    def _handle_project_update_fund_info(self):
        payload = self._read_json()
        project_id = payload.get("project_id")
        if not project_id: raise ValueError("Project id required")
        project = load_project(project_id)
        project["project_name"] = payload.get("project_name") or project.get("project_name") or "Untitled project"
        if payload.get("fund_infos") is not None:
            project["fund_infos"] = payload.get("fund_infos") or {}
        if payload.get("memo_file_path") is not None:
            project["memo_file_path"] = payload.get("memo_file_path") or ""
        memo_path = (project.get("memo_file_path") or "").strip().strip('"').strip("'")
        if not memo_path:
            raise ValueError("Point to a memo file first.")
        if not Path(memo_path).exists():
            raise ValueError(f"Memo file not found: {memo_path}")
        payload_funds = self._normalize_funds_payload(payload.get("funds"))
        project["funds"] = self._persist_project_funds(project_id, payload_funds, project)
        saved = save_project(project)
        result = compute_project_exposure(payload_funds, UPLOADS,
                    parse_mapping_rules(project.get("rules") or ""),
                    issuer_aliases=project.get("issuer_aliases") or "",
                    force_splits=project.get("force_splits") or "")
        try:
            from memo_export import update_sections_in_file
            update_sections_in_file(memo_path, result, sections=("fund_info",), project=saved)
        except PermissionError:
            raise ValueError("Memo is open in Word — close it and try again.")
        self._send_json(200, {"ok": True, "memo_file_path": memo_path})

    def _handle_model_outputs_update(self):
        import model_reader
        payload = self._read_json()
        project_id = payload.get("project_id")
        if not project_id:
            raise ValueError("Project id is required")
        project = load_project(project_id)
        if payload.get("memo_file_path") is not None:
            project["memo_file_path"] = payload.get("memo_file_path") or ""
        if payload.get("model_xlsx_path") is not None:
            project["model_xlsx_path"] = payload.get("model_xlsx_path") or ""
        memo_path = (project.get("memo_file_path") or "").strip().strip('"').strip("'")
        if not memo_path:
            raise ValueError("Point to a memo file first.")
        if not Path(memo_path).exists():
            raise ValueError(f"Memo file not found: {memo_path}")
        xlsx_path = (project.get("model_xlsx_path") or "").strip().strip('"').strip("'")
        if not xlsx_path:
            raise ValueError("Enter the Excel model filepath first.")
        if not Path(xlsx_path).exists():
            raise ValueError(f"Model file not found: {xlsx_path}")
        saved = save_project(project)
        try:
            model_outputs = model_reader.read_model_outputs(xlsx_path)
            update_sections_in_file(
                memo_path,
                {"_model_outputs": model_outputs},
                sections=("model_outputs",),
                project=saved,
            )
        except PermissionError:
            raise ValueError("Could not write the memo — is it open in Word? Close it and try again.")
        self._send_json(200, {"ok": True, "memo_file_path": memo_path})

    def _handle_download_positions_excel(self):
        import io
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        payload = self._read_json()
        result = payload.get("result") or {}
        project_name = payload.get("project_name") or "Project"

        fund_profiles = result.get("fund_profiles") or []
        top_positions = result.get("top_positions") or []
        total_bid = float(result.get("total_bid") or 0)

        fund_names = [fp.get("fund_name") or fp.get("filename") or f"Fund {i+1}"
                      for i, fp in enumerate(fund_profiles)]
        fund_bids = [float(fp.get("bid_amount") or 0) for fp in fund_profiles]
        fund_weights = [float(fp.get("weight") or 0) for fp in fund_profiles]

        # Per-fund position lookup: {issuer_key: fund_level_%}
        # pos.value = project_share = fund_weight × fund_level_%
        # so fund_level_% = pos.value / fund_weight
        fund_pos = []
        for fp in fund_profiles:
            w = float(fp.get("weight") or 0)
            m = {}
            for pos in (fp.get("position_exposure") or []):
                key = str(pos.get("key") or pos.get("label") or "").lower()
                val = float(pos.get("value") or 0)
                m[key] = (val / w) if w > 0 else 0.0
            fund_pos.append(m)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Position Exposure"

        # Row 1: column headers
        headers = ["Issuer", "Project"] + fund_names
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left")

        # Row 2: bid amounts
        ws.cell(2, 1, "Bid Amount").font = Font(italic=True)
        ws.cell(2, 2, total_bid)
        for j, bid in enumerate(fund_bids):
            ws.cell(2, 3 + j, bid)
        for c in range(2, 2 + len(fund_names) + 1):
            ws.cell(2, c).number_format = "#,##0.0"
            ws.cell(2, c).alignment = Alignment(horizontal="center")

        # Column letters for the fund columns (C, D, ...)
        n_funds = len(fund_profiles)
        first_fund_col = get_column_letter(3)
        last_fund_col = get_column_letter(2 + n_funds) if n_funds else get_column_letter(3)

        # Row 3+: one row per issuer
        for r, pos in enumerate(top_positions):
            row = 3 + r
            label = pos.get("label") or ""
            key = str(pos.get("key") or label).lower()

            ws.cell(row, 1, label).alignment = Alignment(horizontal="left")

            # Project % as a formula: =SUMPRODUCT(fund_pcts × fund_bids) / total_bid
            # = SUMPRODUCT(C{row}:{last}{row}, C$2:{last}$2) / B$2
            if n_funds > 0:
                proj_formula = f"=SUMPRODUCT({first_fund_col}{row}:{last_fund_col}{row},{first_fund_col}$2:{last_fund_col}$2)/B$2"
            else:
                proj_formula = "=0"
            c2 = ws.cell(row, 2, proj_formula)
            c2.number_format = "0.0%"
            c2.alignment = Alignment(horizontal="center")

            # Per-fund % — static values (come from the engine)
            for j, pm in enumerate(fund_pos):
                fp_pct = pm.get(key)
                if fp_pct is not None:
                    c = ws.cell(row, 3 + j, fp_pct)
                    c.number_format = "0.0%"
                    c.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 32
        for j in range(len(fund_names) + 1):
            ws.column_dimensions[get_column_letter(2 + j)].width = 16

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        fname = f"{project_name.replace(' ', '_')}_position_exposure.xlsx"
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def _handle_calculate(self):
        payload = self._read_json()
        funds = payload.get("funds") or []
        rules = parse_mapping_rules(payload.get("normalization_rules") or "")
        result = compute_project_exposure(funds, UPLOADS, rules, issuer_aliases=payload.get("issuer_aliases") or "", force_splits=payload.get("force_splits") or "")
        self._send_json(200, result)

    def _read_json(self) -> Dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length) if length else b"{}"
        return json.loads(raw.decode("utf-8") or "{}")

    def log_message(self, format: str, *args):
        return


def run():
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    url = f"http://{HOST}:{PORT}/"
    print(f"Project Balance Exposure Updater running at {url}")
    threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    run()
